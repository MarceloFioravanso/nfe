import os
import time
import base64
import json
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException, TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import chromedriver_autoinstaller
import re
import random
import logging
from datetime import datetime
import platform
import pandas as pd
from pathlib import Path
from preencher_dados_servico import preencher_dados_servico

# Configuração de log
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename=f'logs/nfse_emissao_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log',
    filemode='w'
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('%(levelname)s: %(message)s')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)
logger = logging.getLogger('NFSe_Automacao')

# Pasta para salvar os screenshots
SCREENSHOTS_FOLDER = "logs/imagens"
# Garante que a pasta existe
os.makedirs(SCREENSHOTS_FOLDER, exist_ok=True)

# Função para salvar screenshots na pasta correta
def salvar_screenshot(driver, nome_arquivo):
    """Salva um screenshot na pasta de imagens com o nome especificado"""
    caminho_completo = os.path.join(SCREENSHOTS_FOLDER, nome_arquivo)
    try:
        driver.save_screenshot(caminho_completo)
        logger.info(f"Screenshot salvo em {caminho_completo}")
    except Exception as e:
        logger.error(f"Erro ao salvar screenshot {nome_arquivo}: {e}")

# Instala automaticamente o ChromeDriver compatível
logger.info("Verificando e instalando ChromeDriver compatível...")
chromedriver_autoinstaller.install()

# Carrega .env
logger.info("Carregando variáveis de ambiente do arquivo .env...")
load_dotenv(dotenv_path=".env")

# Verifica se as variáveis foram carregadas corretamente
NFS_URL = os.getenv("NFS_URL", "https://nfse-cachoeirinha.atende.net/autoatendimento/servicos/nfse?redirected=1")
CPF_CNPJ = os.getenv("CPF_CNPJ")
SENHA = os.getenv("SENHA")
PAGINA_DESTINO = "https://nfse-cachoeirinha.atende.net/?rot=1&aca=1#!/sistema/66"

# Caminho do arquivo Excel com as informações das notas fiscais
EXCEL_PATH = r"C:\Users\pesqu\OneDrive\LAF\Adm_Fioravanso\Planejamentos_Controles\Financeiro\_Controle NotaFiscal.xlsx"

# Verifica se as variáveis estão definidas
if not NFS_URL:
    logger.error("ERRO: A variável NFS_URL não está definida no arquivo .env")
    logger.info("Por favor, crie ou edite o arquivo .env com o seguinte conteúdo:")
    logger.info("NFS_URL=https://nfse-cachoeirinha.atende.net/autoatendimento/servicos/nfse?redirected=1")
    logger.info("CPF_CNPJ=seu_cpf_ou_cnpj")
    logger.info("SENHA=sua_senha")
    exit(1)

if not CPF_CNPJ or not SENHA:
    logger.warning("AVISO: Credenciais (CPF_CNPJ e/ou SENHA) não definidas no arquivo .env")
    # Perguntar interativamente se não estiverem definidas
    if not CPF_CNPJ:
        CPF_CNPJ = input("Por favor, digite seu CPF ou CNPJ: ")
    if not SENHA:
        import getpass
        SENHA = getpass.getpass("Por favor, digite sua senha: ")

logger.info(f"URL configurada: {NFS_URL}")
logger.info(f"CPF/CNPJ configurado: {CPF_CNPJ[:4]}{'*' * (len(CPF_CNPJ) - 4) if len(CPF_CNPJ) > 4 else '*****'}")
logger.info(f"Senha configurada: {'*' * 8}")

# Configura o salvamento de arquivos HTML (True para salvar, False para desabilitar)
SALVAR_HTML = True

def salvar_html(driver, nome_arquivo):
    """Função auxiliar para salvar o HTML da página atual"""
    if SALVAR_HTML:
        try:
            with open(f"logs/html/{nome_arquivo}.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            logger.info(f"HTML da página salvo em logs/html/{nome_arquivo}.html")
        except Exception as e:
            logger.error(f"Erro ao salvar HTML: {e}")

def simular_digitacao_humana(elemento, texto):
    """Simula digitação humana inserindo caracteres um a um com tempos variáveis entre eles"""
    # Limpa o campo primeiro
    elemento.clear()
    
    # Velocidades de digitação variáveis
    for caractere in texto:
        elemento.send_keys(caractere)
        # Pequena pausa aleatória para simular digitação real
        time.sleep(random.uniform(0.05, 0.15))
    
    # Pausa final após completar a digitação
    time.sleep(random.uniform(0.2, 0.5))

def verificar_mensagens_erro(driver):
    """
    Verifica se há mensagens de erro na página e retorna uma lista com as mensagens encontradas.
    """
    mensagens = []
    
    # Padrões de texto que podem indicar erro
    padroes_erro = [
        "incorreto", "inválido", "erro", "falha", "não foi possível", 
        "credenciais", "não encontrado", "bloqueado", "expirado", "não autorizado"
    ]
    
    # Busca por elementos que contenham esses textos
    for padrao in padroes_erro:
        try:
            elementos = driver.find_elements(
                By.XPATH, 
                f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{padrao}')]"
            )
            for elem in elementos:
                try:
                    texto = elem.text.strip()
                    if texto and len(texto) > 3 and texto not in mensagens:
                        mensagens.append(texto)
                except:
                    pass
        except:
            continue
    
    # Busca por elementos de alerta/erro específicos
    seletores_erro = [
        ".alert", ".error", ".mensagem-erro", ".aviso", 
        "[class*='alert']", "[class*='error']", "[class*='erro']", 
        "[role='alert']", ".toast-error", ".toast-warning"
    ]
    
    for seletor in seletores_erro:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            for elem in elementos:
                try:
                    texto = elem.text.strip()
                    if texto and len(texto) > 3 and texto not in mensagens:
                        mensagens.append(texto)
                except:
                    pass
        except:
            continue
    
    return mensagens

def verificar_login_sucesso(driver):
    """
    Verifica se o login foi bem-sucedido usando vários métodos.
    Retorna True se o login for bem-sucedido, False caso contrário.
    """
    # Método 1: Verificar se há elementos que indicam que estamos logados
    logger.info("Verificando se o login foi bem-sucedido...")
    
    elementos_pos_login = [
        "a.link_acesso_fiscal",  # Link específico após login
        "a[onclick*='onClickAcessoFiscal']",  # Botão com função específica
        ".menu-logado",  # Menu para usuário logado
        ".user-info",  # Informações do usuário
        ".logout",  # Botão de logout
        ".sair",  # Botão de sair
        "[title='Sair']",  # Elemento com título "Sair"
        "a[href*='logout']",  # Link para logout
        ".botao_cor_tema",  # Botão com tema colorido (geralmente disponível após login)
        ".area-logada"  # Área de usuário logado
    ]
    
    for seletor in elementos_pos_login:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            if elementos:
                logger.info(f"Elemento de login bem-sucedido encontrado: {seletor}")
                return True
        except:
            continue
    
    # Método 2: Verificar pelo texto na página
    textos_pos_login = [
        "Bem-vindo", "Olá", "logado como", "minha conta", 
        "área do contribuinte", "serviços disponíveis", "acessar", "emitir nota"
    ]
    
    for texto in textos_pos_login:
        try:
            elementos = driver.find_elements(
                By.XPATH, 
                f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{texto}')]"
            )
            if elementos:
                logger.info(f"Texto indicativo de login bem-sucedido encontrado: '{texto}'")
                return True
        except:
            continue
    
    # Método 3: Verificar se há mensagens de erro
    mensagens_erro = verificar_mensagens_erro(driver)
    if mensagens_erro:
        logger.error("Mensagens de erro encontradas, login provavelmente falhou:")
        for msg in mensagens_erro:
            logger.error(f"- {msg}")
        return False
    
    # Se não encontrou elementos de sucesso nem elementos de erro, é inconclusivo
    logger.warning("Não foi possível determinar com certeza se o login foi bem-sucedido.")
    return False

def esperar_pagina_carregar(driver, timeout=30):
    """
    Aguarda a página carregar completamente usando diferentes técnicas
    """
    logger.info("Aguardando carregamento completo da página...")
    start_time = time.time()
    
    # Método 1: Aguardar document.readyState
    try:
        wait_curto = WebDriverWait(driver, min(timeout, 10))
        wait_curto.until(lambda d: d.execute_script('return document.readyState') == 'complete')
    except:
        logger.warning("Tempo esgotado aguardando document.readyState == 'complete'")
    
    # Método 2: Aguardar jQuery (se existir)
    try:
        jquery_present = driver.execute_script("return typeof jQuery !== 'undefined'")
        if jquery_present:
            wait_curto = WebDriverWait(driver, min(timeout, 5))
            wait_curto.until(lambda d: d.execute_script('return jQuery.active') == 0)
    except:
        logger.warning("Tempo esgotado aguardando jQuery.active == 0 ou jQuery não encontrado")
    
    # Método 3: Aguardar AJAX (se aplicável)
    # Pausa genérica para permitir processamento AJAX
    remaining_time = timeout - (time.time() - start_time)
    if remaining_time > 0:
        time.sleep(min(remaining_time, 2))
    
    elapsed = time.time() - start_time
    logger.info(f"Página carregada em {elapsed:.2f} segundos")

def clicar_acessar_fiscal(driver):
    """
    Função específica para clicar no botão 'Acessar' após o login.
    Tenta diferentes abordagens para encontrar e clicar no botão.
    """
    logger.info("Procurando o botão 'Acessar' após o login...")
    
    # Lista de seletores para o botão 'Acessar' em ordem de especificidade
    seletores_botao = [
        # Específicos para NFSe Cachoeirinha
        "a.link_acesso_fiscal",
        "a[onclick*='onClickAcessoFiscal']",
        "a.botao_cor_tema",
        
        # Mais genéricos
        ".botao_cor_tema",
        "a.btn-primary",
        ".botao-acesso",
        "[title='Acessar']"
    ]
    
    # Tenta cada seletor
    for seletor in seletores_botao:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            
            if elementos:
                logger.info(f"Botão 'Acessar' encontrado com seletor: {seletor}")
                
                # Tenta tornar o elemento visível se estiver oculto
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elementos[0])
                time.sleep(0.5)
                
                # Tenta métodos diferentes de clique
                try:
                    logger.info("Tentando clicar diretamente...")
                    elementos[0].click()
                    logger.info("Clique direto realizado com sucesso")
                    return True
                except Exception as e:
                    logger.warning(f"Clique direto falhou: {e}")
                    
                    try:
                        logger.info("Tentando clicar via JavaScript...")
                        driver.execute_script("arguments[0].click();", elementos[0])
                        logger.info("Clique via JavaScript realizado com sucesso")
                        return True
                    except Exception as e2:
                        logger.warning(f"Clique via JavaScript falhou: {e2}")
                        
                        # Último recurso - executa função onclick diretamente
                        try:
                            onclick = elementos[0].get_attribute("onclick")
                            if onclick:
                                logger.info(f"Tentando executar onclick: {onclick}")
                                driver.execute_script(onclick)
                                logger.info("Execução de função onclick bem-sucedida")
                                return True
                        except Exception as e3:
                            logger.warning(f"Execução de onclick falhou: {e3}")
                
                # Se chegou aqui, todas as tentativas de clique falharam para este elemento
                logger.warning("Todas as tentativas de clique falharam para este seletor, tentando o próximo...")
        
        except Exception as e:
            logger.warning(f"Erro ao procurar/interagir com seletor {seletor}: {e}")
    
    # Tenta encontrar por texto também
    try:
        elementos_texto = driver.find_elements(By.XPATH, "//a[contains(text(), 'Acessar')] | //button[contains(text(), 'Acessar')]")
        if elementos_texto:
            logger.info("Botão 'Acessar' encontrado por texto")
            elementos_texto[0].click()
            return True
    except Exception as e:
        logger.warning(f"Erro ao buscar botão por texto: {e}")
    
    # Se chegou aqui, nenhum seletor funcionou
    logger.error("Não foi possível encontrar ou clicar no botão 'Acessar'")
    return False

def aguardar_pagina_destino(driver, url_destino, tempo_maximo=300, intervalo=5):
    """
    Aguarda até que a página de destino seja carregada, ou até atingir o tempo máximo.
    
    Args:
        driver: WebDriver do Selenium
        url_destino: URL que estamos esperando carregar
        tempo_maximo: Tempo máximo de espera em segundos (padrão: 5 minutos)
        intervalo: Intervalo entre verificações em segundos
        
    Returns:
        bool: True se a página de destino foi carregada, False caso contrário
    """
    logger.info(f"Aguardando navegação para a página de destino: {url_destino}")
    logger.info(f"Timeout configurado: {tempo_maximo} segundos")
    
    tempo_inicio = time.time()
    url_atual = driver.current_url
    
    while time.time() - tempo_inicio < tempo_maximo:
        # Verifica se a URL atual contém a URL de destino
        url_atual = driver.current_url
        logger.info(f"URL atual: {url_atual}")
        
        if url_destino in url_atual:
            logger.info(f"Página de destino alcançada após {time.time() - tempo_inicio:.2f} segundos")
            return True
        
        # Verifica todas as guias abertas
        for handle in driver.window_handles:
            # Alterna para esta guia
            driver.switch_to.window(handle)
            # Verifica se esta guia tem a URL de destino
            if url_destino in driver.current_url:
                logger.info(f"Página de destino encontrada em uma nova guia após {time.time() - tempo_inicio:.2f} segundos")
                return True
        
        # Volta para a guia original se houver múltiplas guias
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[0])
        
        # Exibe tempo já decorrido
        tempo_decorrido = time.time() - tempo_inicio
        tempo_restante = tempo_maximo - tempo_decorrido
        
        if int(tempo_decorrido) % 30 == 0:  # Log a cada 30 segundos
            logger.info(f"Aguardando... {tempo_decorrido:.0f}s decorridos, {tempo_restante:.0f}s restantes")
            # Tire um screenshot a cada 30 segundos
            salvar_screenshot(driver, f"aguardando_{int(tempo_decorrido)}s.png")
        
        # Pausa antes da próxima verificação
        time.sleep(intervalo)
    
    logger.error(f"Tempo esgotado ({tempo_maximo}s) aguardando a página de destino")
    logger.error(f"URL atual: {url_atual}")
    logger.error(f"URL esperada: {url_destino}")
    return False

def fechar_aviso(driver):
    """
    Tenta fechar o aviso na página de destino clicando no botão Fechar.
    """
    logger.info("Procurando botão 'Fechar' para fechar o aviso...")
    
    # Lista de seletores possíveis para o botão Fechar
    seletores_fechar = [
        "button[name='fechar']",
        "button.botao[name='fechar']",
        "button.__estrutura_componente_base.botao.botao-com-variante",
        "button[myaccesskey='f']",
        "button[zn_id='8']",
        ".botao-com-variante[name='fechar']"
    ]
    
    for seletor in seletores_fechar:
        try:
            # Espera pelo botão ficar visível e clicável
            wait = WebDriverWait(driver, 10)
            fechar_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, seletor)))
            
            logger.info(f"Botão 'Fechar' encontrado com seletor: {seletor}")
            salvar_screenshot(driver, "antes_fechar_aviso.png")
            
            # Tenta clicar no botão
            fechar_btn.click()
            logger.info("Botão 'Fechar' clicado com sucesso")
            time.sleep(2)  # Pequena pausa para o aviso fechar
            salvar_screenshot(driver, "apos_fechar_aviso.png")
            return True
        except Exception as e:
            logger.warning(f"Tentativa de clicar no botão 'Fechar' com seletor {seletor} falhou: {e}")
    
    # Tenta encontrar por texto também
    try:
        elementos_texto = driver.find_elements(By.XPATH, "//button[contains(text(), 'Fechar')]")
        if elementos_texto:
            logger.info("Botão 'Fechar' encontrado por texto")
            elementos_texto[0].click()
            logger.info("Botão 'Fechar' clicado com sucesso")
            time.sleep(2)
            return True
    except Exception as e:
        logger.warning(f"Erro ao buscar botão 'Fechar' por texto: {e}")
    
    # Última tentativa - usar JavaScript para clicar em qualquer botão com nome "fechar"
    try:
        driver.execute_script("document.querySelector('button[name=\"fechar\"]').click();")
        logger.info("Botão 'Fechar' clicado via JavaScript")
        time.sleep(2)
        return True
    except Exception as e:
        logger.warning(f"Clique via JavaScript no botão 'Fechar' falhou: {e}")
    
    logger.error("Não foi possível encontrar ou clicar no botão 'Fechar'")
    return False

def realizar_login(driver, cpf_cnpj, senha):
    """
    Realiza o login no sistema NFSe
    """
    logger.info("Iniciando processo de login...")
    
    # Localiza os elementos de login
    try:
        # Tenta encontrar os campos por vários seletores, começando pelos mais específicos
        logger.info("Tentando localizar elementos do formulário de login...")
        
        # Primeiro tenta seletores específicos para o sistema NFSe de Cachoeirinha
        try:
            wait = WebDriverWait(driver, 10)
            cpf_input = wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input[name='login_usuario']")
            ))
            senha_input = driver.find_element(By.CSS_SELECTOR, "input[name='senha_usuario']")
            login_btn = driver.find_element(By.CSS_SELECTOR, "button.btn-login")
            logger.info("Elementos de login encontrados com seletores específicos para NFSe Cachoeirinha")
        except Exception as e:
            logger.warning(f"Não foi possível encontrar elementos com seletores específicos: {e}")
            
            # Tenta com seletores mais genéricos
            try:
                wait = WebDriverWait(driver, 10)
                cpf_input = wait.until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "input[type='text'][id*='login'], input[type='text'][name*='login'], input[type='text'][id*='cpf'], input[type='text'][name*='cpf'], input[type='text']")
                ))
                senha_input = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                login_btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit'], .btn-login, #btnLogin")
                logger.info("Elementos de login encontrados com seletores genéricos")
            except Exception as e2:
                logger.error(f"Não foi possível encontrar elementos de login: {e2}")
                return False
        
        # Aguarda até que os elementos estejam interagíveis
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, cpf_input.tag_name + "[name='" + cpf_input.get_attribute("name") + "']")))
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, senha_input.tag_name + "[name='" + senha_input.get_attribute("name") + "']")))
        
        logger.info("Preenchendo CPF/CNPJ...")
        cpf_input.click()
        time.sleep(random.uniform(0.1, 0.3))
        simular_digitacao_humana(cpf_input, cpf_cnpj)
        
        logger.info("Preenchendo senha...")
        senha_input.click()
        time.sleep(random.uniform(0.1, 0.3))
        simular_digitacao_humana(senha_input, senha)
        
        # Captura estado antes do login
        salvar_screenshot(driver, "antes_login.png")
        logger.info("Screenshot antes do login salvo")
        
        logger.info("Clicando no botão de login...")
        login_btn.click()
        logger.info("Botão de login clicado")
        
        # Captura estado após clicar
        salvar_screenshot(driver, "apos_clique_login.png")
        
        # Aguarda processamento do login
        esperar_pagina_carregar(driver, timeout=20)
        time.sleep(5)  # Aguarda um pouco mais
        
        # Captura estado após processamento
        salvar_screenshot(driver, "apos_login.png")
        logger.info("Screenshot após processamento de login salvo")
        
        # Salva HTML para análise
        salvar_html(driver, "pagina_apos_login")
        
        # Verifica se o login foi bem-sucedido
        return verificar_login_sucesso(driver)
        
    except Exception as e:
        logger.error(f"Erro durante o processo de login: {e}")
        return False

def clicar_emitir_nota_fiscal(driver):
    """
    Função para clicar no botão 'Emitir Nota Fiscal' após o login.
    Tenta diferentes abordagens para encontrar e clicar no botão.
    """
    logger.info("Procurando o botão 'Emitir Nota Fiscal'...")
    
    # Lista de seletores para o botão 'Emitir Nota Fiscal' em ordem de especificidade
    seletores_botao = [
        # Seletores específicos
        "a[onclick*='emitirNota']",
        "button[name='emitirNota']",
        "a[title='Emitir Nota Fiscal']",
        "button[title='Emitir Nota Fiscal']",
        "a.botao_cor_tema[onclick*='emitir']",
        "a.botao-emitir-nota",
        
        # Mais genéricos
        ".botao_cor_tema",
        ".botao-principal",
        "a.btn-primary",
        ".botao-emitir"
    ]
    
    # Tenta cada seletor
    for seletor in seletores_botao:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            
            if elementos:
                logger.info(f"Botão 'Emitir Nota Fiscal' encontrado com seletor: {seletor}")
                
                # Torna o elemento visível se estiver oculto
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elementos[0])
                time.sleep(0.5)
                
                salvar_screenshot(driver, "antes_emitir_nota_fiscal.png")
                
                # Tenta métodos diferentes de clique
                try:
                    logger.info("Tentando clicar diretamente...")
                    elementos[0].click()
                    logger.info("Clique direto realizado com sucesso")
                    
                    time.sleep(2)  # Pequena pausa após o clique
                    esperar_pagina_carregar(driver, timeout=20)
                    
                    salvar_screenshot(driver, "apos_emitir_nota_fiscal.png")
                    return True
                except Exception as e:
                    logger.warning(f"Clique direto falhou: {e}")
                    
                    try:
                        logger.info("Tentando clicar via JavaScript...")
                        driver.execute_script("arguments[0].click();", elementos[0])
                        logger.info("Clique via JavaScript realizado com sucesso")
                        
                        time.sleep(2)
                        esperar_pagina_carregar(driver, timeout=20)
                        
                        salvar_screenshot(driver, "apos_emitir_nota_fiscal_js.png")
                        return True
                    except Exception as e2:
                        logger.warning(f"Clique via JavaScript falhou: {e2}")
                        
                        # Último recurso - executa função onclick diretamente
                        try:
                            onclick = elementos[0].get_attribute("onclick")
                            if onclick:
                                logger.info(f"Tentando executar onclick: {onclick}")
                                driver.execute_script(onclick)
                                logger.info("Execução de função onclick bem-sucedida")
                                
                                time.sleep(2)
                                esperar_pagina_carregar(driver, timeout=20)
                                
                                salvar_screenshot(driver, "apos_emitir_nota_fiscal_onclick.png")
                                return True
                        except Exception as e3:
                            logger.warning(f"Execução de onclick falhou: {e3}")
                
                # Se chegou aqui, todas as tentativas de clique falharam para este elemento
                logger.warning("Todas as tentativas de clique falharam para este seletor, tentando o próximo...")
        
        except Exception as e:
            logger.warning(f"Erro ao procurar/interagir com seletor {seletor}: {e}")
    
    # Tenta encontrar por texto também
    try:
        xpaths = [
            "//a[contains(text(), 'Emitir Nota')]",
            "//a[contains(text(), 'Emitir') and contains(text(), 'Fiscal')]",
            "//button[contains(text(), 'Emitir Nota')]",
            "//span[contains(text(), 'Emitir Nota')]/..",
            "//div[contains(text(), 'Emitir Nota')]/.."
        ]
        
        for xpath in xpaths:
            elementos_texto = driver.find_elements(By.XPATH, xpath)
            if elementos_texto:
                logger.info(f"Botão 'Emitir Nota Fiscal' encontrado por xpath: {xpath}")
                
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elementos_texto[0])
                time.sleep(0.5)
                
                elementos_texto[0].click()
                logger.info("Botão 'Emitir Nota Fiscal' clicado com sucesso")
                
                time.sleep(2)
                esperar_pagina_carregar(driver, timeout=20)
                
                salvar_screenshot(driver, "apos_emitir_nota_fiscal_xpath.png")
                return True
    except Exception as e:
        logger.warning(f"Erro ao buscar botão por texto: {e}")
    
    # Se chegou aqui, nenhum seletor funcionou
    logger.error("Não foi possível encontrar ou clicar no botão 'Emitir Nota Fiscal'")
    salvar_screenshot(driver, "falha_emitir_nota_fiscal.png")
    salvar_html(driver, "pagina_falha_emitir_nota")
    return False

def clicar_proximo(driver):
    """
    Função para clicar no botão 'Próximo' após selecionar emitir nota fiscal.
    Tenta diferentes abordagens para encontrar e clicar no botão.
    """
    logger.info("Procurando o botão 'Próximo'...")
    
    # Lista de seletores para o botão 'Próximo'
    seletores_botao = [
        # Seletores específicos
        "button[name='proximo']",
        "button[id*='proximo']",
        "button[title='Próximo']",
        "a[title='Próximo']",
        "a.botao_cor_tema[onclick*='proximo']",
        "button.botao-proximo",
        
        # Seletores mais genéricos
        ".botao_cor_tema",
        ".botao-principal",
        ".botao-proximo",
        "button.btn-primary",
        "a.btn-primary"
    ]
    
    # Tenta cada seletor
    for seletor in seletores_botao:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            
            if elementos:
                logger.info(f"Botão 'Próximo' encontrado com seletor: {seletor}")
                
                # Torna o elemento visível se estiver oculto
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elementos[0])
                time.sleep(0.5)
                
                salvar_screenshot(driver, "antes_proximo.png")
                
                # Tenta métodos diferentes de clique
                try:
                    logger.info("Tentando clicar diretamente...")
                    elementos[0].click()
                    logger.info("Clique direto realizado com sucesso")
                    
                    time.sleep(2)  # Pequena pausa após o clique
                    esperar_pagina_carregar(driver, timeout=20)
                    
                    salvar_screenshot(driver, "apos_proximo.png")
                    return True
                except Exception as e:
                    logger.warning(f"Clique direto falhou: {e}")
                    
                    try:
                        logger.info("Tentando clicar via JavaScript...")
                        driver.execute_script("arguments[0].click();", elementos[0])
                        logger.info("Clique via JavaScript realizado com sucesso")
                        
                        time.sleep(2)
                        esperar_pagina_carregar(driver, timeout=20)
                        
                        salvar_screenshot(driver, "apos_proximo_js.png")
                        return True
                    except Exception as e2:
                        logger.warning(f"Clique via JavaScript falhou: {e2}")
                        
                        # Último recurso - executa função onclick diretamente
                        try:
                            onclick = elementos[0].get_attribute("onclick")
                            if onclick:
                                logger.info(f"Tentando executar onclick: {onclick}")
                                driver.execute_script(onclick)
                                logger.info("Execução de função onclick bem-sucedida")
                                
                                time.sleep(2)
                                esperar_pagina_carregar(driver, timeout=20)
                                
                                salvar_screenshot(driver, "apos_proximo_onclick.png")
                                return True
                        except Exception as e3:
                            logger.warning(f"Execução de onclick falhou: {e3}")
                
                # Se chegou aqui, todas as tentativas de clique falharam para este elemento
                logger.warning("Todas as tentativas de clique falharam para este seletor, tentando o próximo...")
        
        except Exception as e:
            logger.warning(f"Erro ao procurar/interagir com seletor {seletor}: {e}")
    
    # Tenta encontrar por texto também
    try:
        xpaths = [
            "//button[contains(text(), 'Próximo')]",
            "//a[contains(text(), 'Próximo')]",
            "//span[contains(text(), 'Próximo')]/parent::button",
            "//div[contains(text(), 'Próximo')]/..",
            "//button[contains(text(), 'Continuar')]",
            "//button[contains(text(), 'Avançar')]"
        ]
        
        for xpath in xpaths:
            elementos_texto = driver.find_elements(By.XPATH, xpath)
            if elementos_texto:
                logger.info(f"Botão 'Próximo' encontrado por xpath: {xpath}")
                
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elementos_texto[0])
                time.sleep(0.5)
                
                elementos_texto[0].click()
                logger.info("Botão 'Próximo' clicado com sucesso")
                
                time.sleep(2)
                esperar_pagina_carregar(driver, timeout=20)
                
                salvar_screenshot(driver, "apos_proximo_xpath.png")
                return True
    except Exception as e:
        logger.warning(f"Erro ao buscar botão 'Próximo' por texto: {e}")
    
    # Se chegou aqui, nenhum seletor funcionou
    logger.error("Não foi possível encontrar ou clicar no botão 'Próximo'")
    salvar_screenshot(driver, "falha_proximo.png")
    salvar_html(driver, "pagina_falha_proximo")
    return False

def verificar_emissao_iniciada(driver):
    """
    Verifica se o fluxo de emissão de nota fiscal foi iniciado com sucesso.
    Procura por elementos que confirmem que estamos na página de emissão.
    
    Returns:
        bool: True se o fluxo foi iniciado com sucesso, False caso contrário
    """
    logger.info("Verificando se o fluxo de emissão foi iniciado com sucesso...")
    
    # Lista de seletores para elementos que podem confirmar que estamos na página de emissão
    seletores_confirmacao = [
        # Campos típicos do formulário de emissão
        "input[name*='tomador']",
        "select[name*='servico']",
        "input[name*='valor']",
        "textarea[name*='descricao']",
        
        # Títulos ou cabeçalhos que podem estar presentes
        "h1, h2, h3, h4",
        ".titulo-pagina",
        ".cabecalho-emissao"
    ]
    
    # Verifica cada seletor
    for seletor in seletores_confirmacao:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            for elemento in elementos:
                texto = elemento.text.lower() if elemento.text else ""
                # Se for um campo de input, verifica o placeholder ou o name
                if elemento.tag_name in ["input", "select", "textarea"]:
                    atributos = [
                        elemento.get_attribute("placeholder") or "",
                        elemento.get_attribute("name") or "",
                        elemento.get_attribute("id") or ""
                    ]
                    # Converte para minúsculas e junta
                    texto_atributos = " ".join([a.lower() for a in atributos if a])
                    
                    # Palavras-chave que indicam que estamos na página de emissão
                    keywords = ["tomador", "servico", "valor", "descricao", "emissao", "nota"]
                    
                    if any(keyword in texto_atributos for keyword in keywords):
                        logger.info(f"Elemento de confirmação encontrado: {elemento.tag_name} com atributos: {texto_atributos}")
                        return True
                
                # Se for um elemento de texto, verifica o texto
                elif texto and len(texto) > 3:
                    # Palavras-chave que podem indicar emissão
                    if any(keyword in texto for keyword in ["emissão", "nota fiscal", "serviço", "dados do tomador"]):
                        logger.info(f"Texto de confirmação encontrado: '{texto}'")
                        return True
        except Exception as e:
            logger.warning(f"Erro ao verificar seletor {seletor}: {e}")
    
    # Se não encontrou elementos confirmatórios, verifica a URL
    url_atual = driver.current_url.lower()
    if any(termo in url_atual for termo in ["emissao", "emitir", "nota", "fiscal", "servico"]):
        logger.info(f"URL indica que estamos na página de emissão: {url_atual}")
        return True
    
    # Se não encontrou nada, salva mais dados para análise
    logger.warning("Não foi possível confirmar que o fluxo de emissão foi iniciado")
    salvar_screenshot(driver, "verificacao_emissao.png")
    salvar_html(driver, "pagina_verificacao_emissao")
    
    return False

def selecionar_tipo_tomador(driver, tipo="Pessoa Jurídica"):
    """
    Seleciona o tipo do tomador no formulário de emissão de nota fiscal.
    
    Args:
        driver: WebDriver do Selenium
        tipo: Tipo do tomador a selecionar (padrão: "Pessoa Jurídica")
    
    Returns:
        bool: True se a seleção foi bem-sucedida, False caso contrário
    """
    logger.info(f"Selecionando tipo do tomador: {tipo}")
    
    # Mapeamento de valores baseados no nome
    valores = {
        "Pessoa Física": "1",
        "Pessoa Jurídica": "2",
        "Pessoa Estrangeira": "3"
    }
    
    # Verifica se o tipo solicitado é válido
    if tipo not in valores:
        logger.error(f"Tipo de tomador inválido: {tipo}. Tipos válidos são: {', '.join(valores.keys())}")
        return False
    
    # Lista de seletores possíveis para o campo de seleção
    seletores = [
        "select[name='tipoTomador']",
        "select.campo-lista",
        "select[aria-label='Tipo do Tomador']",
        "select.estrutura_campo_lista"
    ]
    
    # Tenta localizar o campo de seleção
    select_element = None
    for seletor in seletores:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            if elementos:
                logger.info(f"Elemento select encontrado com seletor: {seletor}")
                select_element = elementos[0]
                break
        except Exception as e:
            logger.warning(f"Erro ao procurar elemento com seletor {seletor}: {e}")
    
    # Se não encontrou, tenta um método mais genérico
    if not select_element:
        try:
            # Tenta encontrar qualquer elemento select na página
            elementos = driver.find_elements(By.TAG_NAME, "select")
            if elementos:
                # Verifica os elementos select para encontrar o de tipo de tomador
                for elem in elementos:
                    try:
                        opcoes = elem.find_elements(By.TAG_NAME, "option")
                        textos_opcoes = [opcao.text for opcao in opcoes]
                        
                        # Verifica se este select contém as opções esperadas
                        if "Pessoa Física" in textos_opcoes and "Pessoa Jurídica" in textos_opcoes:
                            logger.info("Elemento select de tipo de tomador encontrado por análise de opções")
                            select_element = elem
                            break
                    except:
                        continue
        except Exception as e:
            logger.warning(f"Erro ao procurar elementos select genéricos: {e}")
    
    # Se ainda não encontrou, loga o erro e retorna False
    if not select_element:
        logger.error("Não foi possível encontrar o campo de seleção do tipo de tomador")
        salvar_screenshot(driver, "erro_select_tipo_tomador.png")
        salvar_html(driver, "pagina_erro_tipo_tomador")
        return False
    
    # Tenta selecionar o valor desejado
    try:
        # Tenta tornar o elemento visível para interação
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_element)
        time.sleep(0.5)
        
        salvar_screenshot(driver, "antes_selecionar_tipo_tomador.png")
        
        # Método 1: Usar a API do Selenium Select
        from selenium.webdriver.support.ui import Select
        select = Select(select_element)
        
        # Tenta selecionar pelo valor
        try:
            select.select_by_value(valores[tipo])
            logger.info(f"Selecionado '{tipo}' pelo valor: {valores[tipo]}")
        except:
            # Se falhar, tenta selecionar pelo texto visível
            try:
                select.select_by_visible_text(tipo)
                logger.info(f"Selecionado '{tipo}' pelo texto visível")
            except:
                # Último recurso: usar JavaScript
                valor = valores[tipo]
                driver.execute_script(f"arguments[0].value = '{valor}'; arguments[0].dispatchEvent(new Event('change'));", select_element)
                logger.info(f"Selecionado '{tipo}' via JavaScript")
        
        time.sleep(1)  # Pequena pausa para processar a seleção
        
        # Verifica se a seleção foi aplicada
        valor_atual = select_element.get_attribute("value")
        logger.info(f"Valor do select após seleção: {valor_atual}")
        
        salvar_screenshot(driver, "apos_selecionar_tipo_tomador.png")
        
        # Verifica se a seleção foi bem-sucedida
        if valor_atual == valores[tipo]:
            logger.info(f"Tipo de tomador '{tipo}' selecionado com sucesso")
            return True
        else:
            logger.warning(f"A seleção do tipo de tomador pode não ter sido aplicada. Valor esperado: {valores[tipo]}, Valor atual: {valor_atual}")
            return False
    
    except Exception as e:
        logger.error(f"Erro ao selecionar tipo de tomador: {e}")
        salvar_screenshot(driver, "erro_selecionar_tipo_tomador.png")
        import traceback
        logger.error(traceback.format_exc())
        return False

def carregar_dados_excel(caminho_excel):
    """
    Carrega os dados do arquivo Excel e retorna um DataFrame.
    
    Args:
        caminho_excel (str): Caminho para o arquivo Excel
        
    Returns:
        pd.DataFrame: DataFrame com os dados do Excel ou None se houver erro
    """
    try:
        logger.info(f"Carregando dados do arquivo Excel: {caminho_excel}")
        
        # Verifica se o arquivo existe
        if not os.path.exists(caminho_excel):
            logger.error(f"Arquivo Excel não encontrado: {caminho_excel}")
            return None
          # Carrega o arquivo Excel
        # Assume que os dados estão na primeira planilha, com cabeçalho na linha 3 (índice 2)
        df = pd.read_excel(caminho_excel, sheet_name=0, header=2)
        
        logger.info(f"Arquivo Excel carregado com sucesso. {len(df)} linhas encontradas.")
        logger.info(f"Colunas disponíveis: {list(df.columns)}")
        
        return df
        
    except Exception as e:
        logger.error(f"Erro ao carregar arquivo Excel: {e}")
        return None

def encontrar_proxima_nota(df):
    """
    Encontra a próxima nota a ser processada (última linha sem número na primeira coluna mas com dados válidos).
    
    Args:
        df (pd.DataFrame): DataFrame com os dados do Excel
        
    Returns:
        dict: Dados da próxima nota a ser processada ou None se não encontrar
    """
    try:
        if df is None or df.empty:
            logger.error("DataFrame está vazio ou é None")
            return None
        
        logger.info("Procurando a próxima nota a ser processada...")
        
        # Assumindo que a primeira coluna contém os números das notas
        primeira_coluna = df.columns[0]
        logger.info(f"Analisando coluna: {primeira_coluna}")
        
        # Encontra a última linha sem número mas que tenha dados válidos
        for idx in reversed(range(len(df))):
            linha = df.iloc[idx]
            numero_nota = linha[primeira_coluna]
            empresa = linha['Empresa - Razão Social']
            cnpj = linha['CNPJ']
            
            # Verifica se o número da nota está vazio
            numero_vazio = pd.isna(numero_nota) or numero_nota == "" or numero_nota is None
            
            # Verifica se a linha tem dados válidos (empresa ou CNPJ)
            tem_dados = not (pd.isna(empresa) and pd.isna(cnpj))
            
            if numero_vazio and tem_dados:
                logger.info(f"Encontrada linha sem número na posição {idx + 1} (linha {idx + 3} do Excel)")
                
                # Retorna os dados desta linha como dicionário
                dados_nota = linha.to_dict()
                
                # Log dos dados encontrados (mascarando informações sensíveis)
                logger.info("Dados da próxima nota a ser processada:")
                for coluna, valor in dados_nota.items():
                    if pd.notna(valor) and valor != "":
                        # Mascarar CNPJ se necessário
                        if 'CNPJ' in coluna and len(str(valor)) > 6:
                            valor_log = f"{str(valor)[:4]}****{str(valor)[-2:]}"
                        else:
                            valor_log = valor
                        logger.info(f"  {coluna}: {valor_log}")
                
                return {
                    'linha_excel': idx + 3,  # +3 porque pandas é 0-based, Excel começa na linha 1, e temos 2 linhas de header
                    'dados': dados_nota
                }
        
        logger.warning("Não foi encontrada nenhuma linha sem número com dados válidos. Todas as notas podem já ter sido processadas.")
        return None
        
    except Exception as e:
        logger.error(f"Erro ao procurar próxima nota: {e}")
        return None

def mapear_dados_nota(dados_nota):
    """
    Mapeia os dados do Excel para as variáveis necessárias para emissão da nota fiscal.
    
    Args:
        dados_nota (dict): Dados da linha do Excel
        
    Returns:
        dict: Dados mapeados para emissão da nota fiscal
    """
    try:
        logger.info("Mapeando dados da nota fiscal...")
        # Mapeamento baseado nas colunas reais do Excel
        # Ajuste os nomes das colunas conforme encontrado no arquivo
        mapeamento = {
            'cnpj_tomador': dados_nota.get('CNPJ', ''),
            'razao_social': dados_nota.get('Empresa - Razão Social', ''),
            'nome_fantasia': dados_nota.get('Empresa - Razão Social', ''),  # Usando o mesmo campo
            'endereco': dados_nota.get('Endereço', ''),
            'numero': dados_nota.get('Número', ''),  # Novo campo para número
            'complemento': dados_nota.get('Complemento', ''),  # Campo para complemento
            'bairro': dados_nota.get('BAIRRO', ''),
            'cidade': dados_nota.get('Município', ''),
            'uf': dados_nota.get('Estado', ''),
            'cep': dados_nota.get('CEP', ''),
            'email': '',  # Não há campo de email específico no Excel
            'telefone': '',  # Não há campo de telefone específico no Excel
            'descricao_servico': dados_nota.get('Descrição', ''),
            'valor_servico': dados_nota.get('Total', 0),
            'aliquota_iss': '',  # Não há campo específico de alíquota
            'valor_iss': '',  # Não há campo específico de ISS
            'data_competencia': '',  # Pode ser construída com Dia/Mês/Ano Emis
            'codigo_servico': '',  # Não há código de serviço específico
            'observacoes': dados_nota.get('Observações', ''),
            'inscricao_estadual': dados_nota.get('Inscrição Estadual', ''),
            'inscricao_municipal': dados_nota.get('Inscrição Municipal', ''),
            'irrf': dados_nota.get('IRRF(1,5%) ou (4,8%)', 0),
            'pis': dados_nota.get('PIS (0,65%)', 0),
            'cofins': dados_nota.get('Cofins (3%)', 0),
            'csll': dados_nota.get('Contr. Social - CSLL (1%)', 0),            'total_impostos': dados_nota.get('Total Impostos', 0),
            'valor_liquido': dados_nota.get('Líquido', 0),
            # Campos adicionais para a descrição do serviço
            'vencimento_dia': dados_nota.get('Dia Venc.', ''),    # Coluna para dia do vencimento
            'vencimento_mes': dados_nota.get('Mês Venc.', ''),    # Coluna para mês do vencimento
            'vencimento_ano': dados_nota.get('Ano Venc.', ''),    # Coluna para ano do vencimento
            'parcela': dados_nota.get('Parcela', '')              # Coluna para número da parcela
        }
          # Validações básicas
        campos_obrigatorios = ['cnpj_tomador', 'razao_social', 'valor_servico', 'descricao_servico']
        campos_faltando = []
        
        for campo in campos_obrigatorios:
            valor = mapeamento.get(campo)
            if not valor or (isinstance(valor, str) and valor.strip() == '') or pd.isna(valor):
                campos_faltando.append(campo)
        
        if campos_faltando:
            logger.warning(f"Campos obrigatórios faltando ou vazios: {', '.join(campos_faltando)}")
        
        # Log dos dados mapeados
        logger.info("Dados mapeados:")
        for campo, valor in mapeamento.items():
            if valor and not pd.isna(valor) and str(valor).strip() != '':
                # Mascarar dados sensíveis
                if 'cnpj' in campo.lower():
                    valor_log = f"{str(valor)[:4]}****{str(valor)[-2:]}" if len(str(valor)) > 6 else "****"
                else:
                    valor_log = valor
                logger.info(f"  {campo}: {valor_log}")
        
        return mapeamento
        
    except Exception as e:
        logger.error(f"Erro ao mapear dados da nota: {e}")
        return None

def atualizar_numero_nota_excel(caminho_excel, linha_excel, numero_nota):
    """
    Atualiza o número da nota fiscal no arquivo Excel após a emissão.
    
    Args:
        caminho_excel (str): Caminho para o arquivo Excel
        linha_excel (int): Número da linha no Excel (1-based, já considerando o cabeçalho)
        numero_nota (str): Número da nota fiscal emitida
        
    Returns:
        bool: True se a atualização foi bem-sucedida, False caso contrário
    """
    try:
        logger.info(f"Atualizando Excel - Linha {linha_excel} com número da nota: {numero_nota}")
        
        # Faz backup do arquivo original antes de qualquer modificação
        backup_path = caminho_excel.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        import shutil
        shutil.copy2(caminho_excel, backup_path)
        logger.info(f"Backup criado: {backup_path}")
        
        # Usa openpyxl para preservar toda a formatação e dados do Excel
        try:
            from openpyxl import load_workbook
            
            # Carrega o arquivo Excel
            logger.info(f"Carregando arquivo Excel: {caminho_excel}")
            book = load_workbook(caminho_excel)
            sheet = book.active
            
            # Atualiza o valor na célula correta (linha Excel é 1-based)
            # A primeira coluna é A(1)
            cell = sheet.cell(row=linha_excel, column=1)
            valor_anterior = cell.value
            cell.value = numero_nota
            
            logger.info(f"Atualizando célula A{linha_excel}: Valor anterior: {valor_anterior} -> Novo valor: {numero_nota}")
            
            # Salva o arquivo atualizado
            book.save(caminho_excel)
            logger.info(f"Arquivo Excel atualizado com sucesso na linha {linha_excel}, coluna A")
            
            return True
            
        except ImportError:
            # Se openpyxl não estiver instalado, usa pandas como fallback
            logger.warning("Biblioteca openpyxl não encontrada. Usando pandas como alternativa.")
            
            # Carrega o arquivo Excel com pandas (pode perder alguma formatação)
            df = pd.read_excel(caminho_excel, sheet_name=0, header=2)
            
            # Calcula o índice do DataFrame (0-based)
            # Ajusta o índice considerando o header=2
            indice_df = linha_excel - 3
            
            if indice_df < 0 or indice_df >= len(df):
                logger.error(f"Linha {linha_excel} está fora do intervalo válido do DataFrame (índice {indice_df})")
                return False
            
            # Registra o valor anterior antes de atualizar
            primeira_coluna = df.columns[0]
            valor_anterior = df.iloc[indice_df, 0]
            
            # Atualiza a primeira coluna
            df.iloc[indice_df, 0] = numero_nota
            logger.info(f"Atualizando valor na linha {linha_excel}: {valor_anterior} -> {numero_nota}")
            
            # Salva o arquivo atualizado, preservando o índice original
            df.to_excel(caminho_excel, sheet_name='Sheet1', index=False, header=True)
            logger.info(f"Arquivo Excel atualizado com pandas na linha {linha_excel}")
            
            return True
        
    except Exception as e:
        logger.error(f"Erro ao atualizar arquivo Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        # Sugere ao usuário atualizar manualmente
        logger.error(f"\nPor favor, atualize manualmente o arquivo Excel:")
        logger.error(f"1. Abra o arquivo: {caminho_excel}")
        logger.error(f"2. Vá até a linha {linha_excel}")
        logger.error(f"3. Na primeira coluna, insira o número da nota: {numero_nota}")
        
        return False

def buscar_empresa_por_cnpj(driver, cnpj, nome_empresa):
    """
    Busca empresa pelo CNPJ e seleciona nos resultados de pesquisa.
    
    Args:
        driver: WebDriver do Selenium
        cnpj: CNPJ da empresa para busca
        nome_empresa: Nome da empresa para confirmar na seleção de resultados
        
    Returns:
        bool: True se a busca e seleção foi bem-sucedida, False caso contrário
    """
    try:
        logger.info(f"Buscando empresa com CNPJ: {cnpj[:4]}****{cnpj[-2:]} - {nome_empresa}")
        wait = WebDriverWait(driver, 10)
        
        # Procura pelo campo de busca de CNPJ/Razão Social
        campo_busca_seletores = [
            'input[name="Tomador.nomeRazao"]',
            'input[placeholder*="Pesquisar por nome ou CNPJ"]',
            'input.campo-texto[aria-label*="campo"]',
            'input.campo-texto',
            'input[name*="Tomador"][name*="nomeRazao"]'
        ]
        
        campo_busca = None
        for seletor in campo_busca_seletores:
            try:
                elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
                if elementos:
                    campo_busca = elementos[0]
                    logger.info(f"Campo de busca de CNPJ encontrado com seletor: {seletor}")
                    break
            except Exception as e:
                logger.warning(f"Erro ao procurar campo com seletor {seletor}: {e}")
        
        if not campo_busca:
            logger.error("Campo de busca de CNPJ/Razão Social não encontrado")
            salvar_screenshot(driver, "erro_campo_busca_cnpj_nao_encontrado.png")
            return False
        
        # Clica no campo de busca
        campo_busca.click()
        time.sleep(0.5)
        
        # Limpa o campo e insere o CNPJ
        campo_busca.clear()
        simular_digitacao_humana(campo_busca, cnpj)
        logger.info(f"CNPJ inserido no campo de busca")
        salvar_screenshot(driver, "apos_inserir_cnpj.png")
        
        # Aguarda os resultados
        time.sleep(2)
        
        # Busca os resultados da pesquisa
        logger.info("Procurando resultados da busca de CNPJ...")
        resultados_seletores = [
            'table.tabela-resultado tr',
            'div.resultado-pesquisa',
            'td[name="nomeRazao"]',
            'tr[onclick*="selecionar"]'
        ]
        
        for seletor in resultados_seletores:
            try:
                elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
                if elementos:
                    for elemento in elementos:
                        texto_elemento = elemento.text.upper() if elemento.text else ""
                        # Verificar se o texto do elemento contém o nome da empresa
                        if nome_empresa.upper() in texto_elemento:
                            logger.info(f"Empresa encontrada nos resultados: {texto_elemento}")
                            salvar_screenshot(driver, "empresa_encontrada.png")
                            
                            # Clica no resultado
                            try:
                                elemento.click()
                                logger.info("Empresa selecionada com sucesso")
                                salvar_screenshot(driver, "apos_selecionar_empresa.png")
                                time.sleep(1)
                                return True
                            except Exception as e:
                                logger.error(f"Erro ao clicar no resultado: {e}")
                                try:
                                    # Tenta clicar via JavaScript
                                    driver.execute_script("arguments[0].click();", elemento)
                                    logger.info("Empresa selecionada via JavaScript")
                                    return True
                                except:
                                    logger.error("Falha ao clicar na empresa usando JavaScript")
                                    return False
            except Exception as e:
                logger.warning(f"Erro ao procurar resultados com seletor {seletor}: {e}")
        
        # Se não encontrou resultados específicos, busca por td com name="nomeRazao"
        try:
            logger.info("Tentando localizar célula com name='nomeRazao'...")
            elementos = driver.find_elements(By.CSS_SELECTOR, 'td[name="nomeRazao"]')
            for elemento in elementos:
                texto_elemento = elemento.text.upper() if elemento.text else ""
                if nome_empresa.upper() in texto_elemento:
                    logger.info(f"Empresa encontrada em td[name='nomeRazao']: {texto_elemento}")
                    elemento.click()
                    logger.info("Empresa selecionada com sucesso")
                    return True
        except Exception as e:
            logger.error(f"Erro ao buscar por td[name='nomeRazao']: {e}")
        
        logger.error(f"Empresa '{nome_empresa}' não encontrada nos resultados da busca")
        salvar_screenshot(driver, "empresa_nao_encontrada.png")
        return False
    
    except Exception as e:
        logger.error(f"Erro ao buscar empresa por CNPJ: {e}")
        salvar_screenshot(driver, "erro_busca_cnpj.png")
        return False

def preencher_dados_tomador(driver, dados_nota):
    """
    Preenche os dados do tomador no formulário com base nos dados do Excel.
    
    Args:
        driver: WebDriver do Selenium
        dados_nota (dict): Dados mapeados da nota fiscal
        
    Returns:
        bool: True se o preenchimento foi bem-sucedido, False caso contrário
    """
    try:
        logger.info("Iniciando preenchimento dos dados do tomador...")
        wait = WebDriverWait(driver, 10)
        
        # Aguarda um pouco para a página carregar após selecionar tipo do tomador
        time.sleep(2)
        
        # Verifica se o checkbox "Endereço Alternativo" está selecionado
        logger.info("Verificando se o checkbox 'Endereço Alternativo' está marcado...")
        endereco_alternativo_selecionado = False
        
        try:
            # Seletores possíveis para o checkbox de endereço alternativo
            seletores_checkbox = [
                'input[name="usaEnderecoAlternativo"]',
                'input[aria-label="Endereço Alternativo"]',
                'input.__estrutura_componente_base.campo.estrutura_check_tipo_toggle[name="usaEnderecoAlternativo"]',
                'input.estrutura_check_tipo_toggle[name="usaEnderecoAlternativo"]',
                'input[type="checkbox"][name*="endereco"]',
                'input[type="checkbox"][aria-label*="endereco"]',
                'input[type="checkbox"][id*="endereco"]',
                'input.__estrutura_componente_base.campo.estrutura_check_tipo_toggle'
            ]
            
            for seletor in seletores_checkbox:
                checkboxes = driver.find_elements(By.CSS_SELECTOR, seletor)
                if checkboxes:
                    checkbox = checkboxes[0]
                    # Verifica usando múltiplos métodos
                    is_selected_attr = checkbox.is_selected()
                    is_checked_attr = checkbox.get_attribute("checked") == "true"
                    has_checked_class = "checked" in (checkbox.get_attribute("class") or "")
                    aria_checked = checkbox.get_attribute("aria-checked") == "true"
                    
                    # Combinação dos resultados (qualquer um positivo indica que está marcado)
                    endereco_alternativo_selecionado = any([is_selected_attr, is_checked_attr, has_checked_class, aria_checked])
                    
                    logger.info(f"Checkbox 'Endereço Alternativo' encontrado com seletor: {seletor}")
                    logger.info(f"Estado do checkbox - is_selected(): {is_selected_attr}")
                    logger.info(f"Estado do checkbox - atributo checked: {is_checked_attr}")
                    logger.info(f"Estado do checkbox - classe checked: {has_checked_class}")
                    logger.info(f"Estado do checkbox - aria-checked: {aria_checked}")
                    logger.info(f"Conclusão: Checkbox 'Endereço Alternativo' está {'MARCADO' if endereco_alternativo_selecionado else 'DESMARCADO'}")
                    break
            
            # Se não encontrou por CSS, tenta por XPath
            if not checkboxes:
                xpath_seletores = [
                    "//input[@name='usaEnderecoAlternativo']",
                    "//input[@aria-label='Endereço Alternativo']",
                    "//label[contains(text(), 'Endereço Alternativo')]/preceding-sibling::input",
                    "//label[contains(text(), 'Endereço Alternativo')]/following-sibling::input",
                    "//label[contains(., 'Endereço Alternativo')]//input",
                    "//span[contains(text(), 'Endereço Alternativo')]/preceding-sibling::input",
                    "//span[contains(text(), 'Endereço Alternativo')]/following-sibling::input",
                    "//input[contains(@name, 'endereco') and @type='checkbox']"
                ]
                
                for xpath in xpath_seletores:
                    elementos = driver.find_elements(By.XPATH, xpath)
                    if elementos:
                        checkbox = elementos[0]
                        # Verifica usando múltiplos métodos
                        is_selected_attr = checkbox.is_selected()
                        is_checked_attr = checkbox.get_attribute("checked") == "true"
                        has_checked_class = "checked" in (checkbox.get_attribute("class") or "")
                        aria_checked = checkbox.get_attribute("aria-checked") == "true"
                        
                        # Combinação dos resultados
                        endereco_alternativo_selecionado = any([is_selected_attr, is_checked_attr, has_checked_class, aria_checked])
                        
                        logger.info(f"Checkbox 'Endereço Alternativo' encontrado com XPath: {xpath}")
                        logger.info(f"Estado do checkbox - is_selected(): {is_selected_attr}")
                        logger.info(f"Estado do checkbox - atributo checked: {is_checked_attr}")
                        logger.info(f"Estado do checkbox - classe checked: {has_checked_class}")
                        logger.info(f"Estado do checkbox - aria-checked: {aria_checked}")
                        logger.info(f"Conclusão: Checkbox 'Endereço Alternativo' está {'MARCADO' if endereco_alternativo_selecionado else 'DESMARCADO'}")
                        break
            
            # Último recurso: verificação visual (captura elementos visuais que parecem checkboxes marcados)
            if not endereco_alternativo_selecionado:
                logger.info("Tentando detecção visual de checkbox marcado...")
                
                # Procura por elementos que parecem visualmente checkboxes marcados
                elementos_visuais = driver.find_elements(By.CSS_SELECTOR, 
                    '.checked, .selected, [class*="checkbox"][class*="checked"], [class*="checkbox"][class*="selected"], '
                    '[class*="check"][class*="active"], [class*="toggle"][class*="active"], [class*="toggle"][class*="on"]')
                
                for elem in elementos_visuais:
                    try:
                        # Procura por texto relacionado a "Endereço Alternativo" próximo ao elemento
                        parent = elem.find_element(By.XPATH, './parent::*')
                        texto_proximo = parent.text;

                        if 'endereço alternativo' in texto_proximo.lower():
                            logger.info("Checkbox 'Endereço Alternativo' encontrado por detecção visual")
                            endereco_alternativo_selecionado = True
                            break

                        # Verifica também elementos próximos
                        elementos_proximos = driver.find_elements(By.XPATH, 
                            f'//label[contains(translate(text(), "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz"), "endereço alternativo")]')

                        if elementos_proximos:
                            # Verifica se o elemento visual está próximo do texto
                            for texto_elem in elementos_proximos:
                                if elem.location['y'] - 50 < texto_elem.location['y'] < elem.location['y'] + 50:
                                    logger.info("Checkbox 'Endereço Alternativo' encontrado por proximidade com texto")
                                    endereco_alternativo_selecionado = True
                                    break

                        if endereco_alternativo_selecionado:
                            break
                    except Exception as e:
                        logger.debug(f"Erro na detecção visual: {e}")
            
            # Salva screenshot do estado do checkbox
            salvar_screenshot(driver, "verificacao_endereco_alternativo.png")
            
            if endereco_alternativo_selecionado:
                logger.info("O checkbox 'Endereço Alternativo' está marcado. Não é necessário preencher os dados de endereço.")
                
                # Clica direto no botão próximo
                logger.info("Tentando clicar no botão 'Próximo'...")
                seletores_proximo = [
                    'button[name="botao_proximo"]',
                    'button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_janela_proximo[name="botao_proximo"][myaccesskey="p"]',
                    'button.botao-com-variante[name="botao_proximo"]',
                    'button.__estrutura_componente_base.botao.botao-com-variante',
                    'button[myaccesskey="p"]',
                    'button.estrutura_botao_janela_proximo',
                    # Seletores adicionais para o botão próximo
                    'button.botao_proximo',
                    'button.botao-com-variante[type="submit"]',
                    'button:not([disabled])[name*="proximo"]',
                    'button:not([disabled])[id*="proximo"]',
                    'button:not([disabled])[title*="Próximo"]',
                    'button:not([disabled])[aria-label*="Próximo"]'
                ]
                
                for seletor in seletores_proximo:
                    try:
                        elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
                        elementos_visiveis = [e for e in elementos if e.is_displayed()]
                        
                        if elementos_visiveis:
                            botao_proximo = elementos_visiveis[0]
                            logger.info(f"Botão 'Próximo' encontrado com seletor: {seletor}")
                            
                            # Torna o botão visível
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao_proximo)
                            time.sleep(0.5)
                            
                            # Salva screenshot antes de clicar
                            salvar_screenshot(driver, "antes_clicar_proximo_endereco_alternativo.png")
                            
                            # Tenta clicar no botão
                            try:
                                botao_proximo.click()
                                logger.info("Botão 'Próximo' clicado com sucesso")
                                time.sleep(2)
                                salvar_screenshot(driver, "apos_clicar_proximo_endereco_alternativo.png")
                                
                                # Verifica se a página mudou após o clique
                                try:
                                    # Procura por elementos que indicam que a página foi alterada
                                    nova_pagina = WebDriverWait(driver, 5).until(
                                        EC.presence_of_element_located((By.CSS_SELECTOR, 
                                            'input[name*="servico"], input[name*="descricao"], textarea[name*="descricao"], '
                                            'input[name*="valor"], input[name*="Valores"]'
                                        ))
                                    )
                                    logger.info("Navegação confirmada: Encontrados elementos da próxima página")
                                except:
                                    logger.warning("Não foi possível confirmar se a navegação ocorreu com sucesso")
                                
                                return True
                            except Exception as e:
                                logger.warning(f"Clique direto falhou, tentando via JavaScript: {e}")
                                try:
                                    driver.execute_script("arguments[0].click();", botao_proximo)
                                    logger.info("Botão 'Próximo' clicado via JavaScript")
                                    time.sleep(2)
                                    salvar_screenshot(driver, "apos_clicar_proximo_endereco_alternativo_js.png")
                                    
                                    # Verifica se a página mudou após o clique via JavaScript
                                    try:
                                        # Procura por elementos que indicam que a página foi alterada
                                        nova_pagina = WebDriverWait(driver, 5).until(
                                            EC.presence_of_element_located((By.CSS_SELECTOR, 
                                                'input[name*="servico"], input[name*="descricao"], textarea[name*="descricao"], '
                                                'input[name*="valor"], input[name*="Valores"]'
                                            ))
                                        )
                                        logger.info("Navegação confirmada: Encontrados elementos da próxima página")
                                    except:
                                        logger.warning("Não foi possível confirmar se a navegação ocorreu com sucesso")
                                    
                                    return True
                                except Exception as e2:
                                    logger.error(f"Clique via JavaScript também falhou: {e2}")
                    except Exception as e:
                        logger.warning(f"Erro ao procurar botão 'Próximo' com seletor {seletor}: {e}")
                
                # Se não encontrou por CSS, tenta por XPath
                xpath_botoes = [
                    "//button[contains(text(), 'Próximo')]",
                    "//a[contains(text(), 'Próximo')]",
                    "//span[contains(text(), 'Próximo')]/parent::button",
                    "//div[contains(text(), 'Próximo')]/..",
                    "//button[contains(text(), 'Continuar')]",
                    "//button[contains(text(), 'Avançar')]"
                ]
                
                for xpath in xpath_botoes:
                    try:
                        elementos = driver.find_elements(By.XPATH, xpath)
                        elementos_visiveis = [e for e in elementos if e.is_displayed()]
                        
                        if elementos_visiveis:
                            botao_proximo = elementos_visiveis[0]
                            logger.info(f"Botão 'Próximo' encontrado com XPath: {xpath}")
                            
                            # Torna o botão visível
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao_proximo)
                            time.sleep(0.5)
                            
                            # Salva screenshot antes de clicar
                            salvar_screenshot(driver, "antes_clicar_proximo_xpath_endereco_alternativo.png")
                            
                            try:
                                botao_proximo.click()
                                logger.info("Botão 'Próximo' clicado com sucesso via XPath")
                                time.sleep(2)
                                salvar_screenshot(driver, "apos_clicar_proximo_xpath_endereco_alternativo.png")
                                return True
                            except Exception as e:
                                logger.warning(f"Clique direto por XPath falhou, tentando via JavaScript: {e}")
                                try:
                                    driver.execute_script("arguments[0].click();", botao_proximo)
                                    logger.info("Botão 'Próximo' clicado via JavaScript (XPath)")
                                    time.sleep(2)
                                    salvar_screenshot(driver, "apos_clicar_proximo_xpath_js_endereco_alternativo.png")
                                    return True
                                except Exception as e2:
                                    logger.error(f"Clique via JavaScript (XPath) também falhou: {e2}")
                    except Exception as e:
                        logger.warning(f"Erro ao procurar botão 'Próximo' com XPath {xpath}: {e}")
                
                logger.error("Não foi possível encontrar ou clicar no botão 'Próximo'")
                salvar_screenshot(driver, "erro_botao_proximo_nao_encontrado_endereco_alternativo.png")
                return False
                
        except Exception as e:
            logger.warning(f"Erro ao verificar checkbox de endereço alternativo: {e}")
            import traceback
            logger.warning(traceback.format_exc())
            # Em caso de erro, continua com o preenchimento normal do endereço
        
        # Após selecionar empresa, agora precisamos preencher os campos de endereço
        logger.info("Preenchendo dados de endereço do tomador...")# Processa o endereço para separar logradouro e número
        endereco_completo = dados_nota.get('endereco', '')
        logradouro = endereco_completo
        numero = dados_nota.get('numero', '')
        complemento = dados_nota.get('complemento', '')
        
        # Importa re para usar expressões regulares
        import re
        
        # Estratégia 1: Se tiver o campo número preenchido no Excel, usa diretamente
        if numero:
            logger.info(f"Utilizando número do campo específico: {numero}")
            # Se o número estiver no endereço completo, remove do endereço para evitar duplicidade
            if str(numero) in endereco_completo:
                logradouro = re.sub(r',?\s*' + re.escape(str(numero)) + r'\s*,?', '', endereco_completo).strip()
        
        # Estratégia 2: Se não tiver número e o endereço contiver uma vírgula, tenta extrair
        elif ',' in endereco_completo:
            partes = endereco_completo.split(',')
            logradouro = partes[0].strip()
            # A segunda parte pode ser o número ou o complemento
            if len(partes) > 1:
                segunda_parte = partes[1].strip()
                # Extrai números da segunda parte
                numeros = re.findall(r'\b\d+\b', segunda_parte)
                if numeros:
                    numero = numeros[0]
                    # O restante pode ser complemento
                    resto = re.sub(r'\b' + re.escape(numero) + r'\b', '', segunda_parte).strip()
                    if resto:
                        complemento = (complemento + ' ' + resto).strip() if complemento else resto
                else:
                    # Se não houver números na segunda parte, considera tudo como complemento
                    complemento = (complemento + ' ' + segunda_parte).strip() if complemento else segunda_parte
        
        # Estratégia 3: Tenta localizar o número no final do endereço (comum em endereços brasileiros)
        elif not numero:
            # Padrão típico: "Rua Nome da Rua 123"
            match = re.search(r'(.*[^\d])\s+(\d+)\s*$', endereco_completo)
            if match:
                logradouro = match.group(1).strip()
                numero = match.group(2)
                
        # Se ainda não encontrou o número e é obrigatório, define um valor padrão
        if not numero:
            numero = "S/N"
            logger.warning(f"Número não encontrado no endereço, usando valor padrão: {numero}")
            
        # Limpa possíveis vírgulas sobrando no logradouro
        logradouro = logradouro.rstrip(',').strip()
        
        logger.info(f"Endereço processado: Logradouro='{logradouro}', Número='{numero}', Complemento='{complemento}'")
          # Lista de campos a serem preenchidos
        campos_endereco = {
           

            'cep': {
                'seletores': ['input[name="InformacoesTomador.cep"]', 'input[aria-label="CEP"]', 'input[name*="cep"]'],
                'valor': dados_nota.get('cep', ''),
                'obrigatorio': True
            },
            'logradouro': {
                'seletores': ['input[name="InformacoesTomador.logradouro"]', 'input[aria-label="Logradouro"]', 'input[name*="logradouro"]'],
                'valor': logradouro,
                'obrigatorio': True
            },
            'numero': {
                'seletores': ['input[name="InformacoesTomador.numero"]', 'input[aria-label="Número"]', 'input[name*="numero"]'],
                'valor': str(numero),
                'obrigatorio': True
            },
            'complemento': {
                'seletores': ['input[name="InformacoesTomador.complemento"]', 'input[aria-label="Complemento"]', 'input[name*="complemento"]'],
                'valor': complemento,
                'obrigatorio': False
            },
            'bairro': {
                'seletores': ['input[name="InformacoesTomador.bairro"]', 'input[aria-label="Bairro"]', 'input[name*="bairro"]', 'input[class*="campo-texto"][aria-label="Bairro"]'],
                'valor': dados_nota.get('bairro', ''),
                'obrigatorio': False
            }
        }
        
        # Tenta preencher cada campo de endereço
        campos_preenchidos = 0
        campos_falharam = []
        
        for nome_campo, config in campos_endereco.items():
            valor = config['valor']
            
            # Se não tiver um valor, e o campo não for obrigatório, pula
            if not valor and not config['obrigatorio']:
                logger.info(f"Campo {nome_campo} não é obrigatório e não tem valor. Pulando...")
                continue
            
            # Tenta cada seletor para encontrar o campo
            campo_encontrado = False
            for seletor in config['seletores']:
                try:
                    elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
                    if elementos:
                        elemento = elementos[0]
                        logger.info(f"Campo {nome_campo} encontrado com seletor: {seletor}")
                        
                        # Torna o elemento visível
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
                        time.sleep(0.5)
                        
                        # Limpa e preenche o campo
                        elemento.clear()
                        simular_digitacao_humana(elemento, valor)
                        logger.info(f"Campo {nome_campo} preenchido com: '{valor}'")
                        
                        campos_preenchidos += 1
                        campo_encontrado = True
                        break
                except Exception as e:
                    logger.warning(f"Erro ao procurar/preencher {nome_campo} com seletor {seletor}: {e}")
            
            if not campo_encontrado and config['obrigatorio']:
                logger.error(f"Não foi possível encontrar campo obrigatório: {nome_campo}")
                campos_falharam.append(nome_campo)
        
        # Se todos os campos obrigatórios foram preenchidos
        if not campos_falharam:
            logger.info(f"Dados de endereço preenchidos com sucesso: {campos_preenchidos} campo(s)")
            salvar_screenshot(driver, "apos_preencher_endereco.png")
            
            # Clica no botão próximo
            logger.info("Tentando clicar no botão 'Próximo'...")
            botao_proximo = None
              # Lista de seletores para o botão próximo
            seletores_proximo = [
                'button[name="botao_proximo"]',
                'button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_janela_proximo[name="botao_proximo"][myaccesskey="p"]',
                'button.botao-com-variante[name="botao_proximo"]',
                'button.__estrutura_componente_base.botao.botao-com-variante',
                'button[myaccesskey="p"]',
                'button.estrutura_botao_janela_proximo'
            ]
            
            # Tenta encontrar e clicar no botão próximo
            for seletor in seletores_proximo:
                try:
                    elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
                    if elementos:
                        botao_proximo = elementos[0]
                        logger.info(f"Botão 'Próximo' encontrado com seletor: {seletor}")
                        
                        # Torna o botão visível
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao_proximo)
                        time.sleep(0.5)
                        
                        # Tenta clicar no botão
                        try:
                            botao_proximo.click()
                            logger.info("Botão 'Próximo' clicado com sucesso")
                            time.sleep(2)
                            salvar_screenshot(driver, "apos_clicar_proximo_endereco.png")
                            return True
                        except Exception as e:
                            logger.warning(f"Clique direto falhou, tentando via JavaScript: {e}")
                            try:
                                driver.execute_script("arguments[0].click();", botao_proximo)
                                logger.info("Botão 'Próximo' clicado via JavaScript")
                                time.sleep(2)
                                salvar_screenshot(driver, "apos_clicar_proximo_endereco_js.png")
                                return True
                            except Exception as e2:
                                logger.error(f"Clique via JavaScript também falhou: {e2}")
                except Exception as e:
                    logger.warning(f"Erro ao procurar botão 'Próximo' com seletor {seletor}: {e}")
            
            # Se chegou aqui, não encontrou o botão próximo
            logger.error("Não foi possível encontrar ou clicar no botão 'Próximo'")
            salvar_screenshot(driver, "erro_botao_proximo_nao_encontrado.png")
            return False
        else:
            logger.error(f"Falha ao preencher campos obrigatórios: {campos_falharam}")
            salvar_screenshot(driver, "erro_preenchimento_campos_endereco.png")
            return False
            
    except Exception as e:
        logger.error(f"Erro durante preenchimento dos dados de endereço do tomador: {e}")
        import traceback
        logger.error(traceback.format_exc())
        salvar_screenshot(driver, "erro_geral_preenchimento_endereco.png")
        return False

def obter_caminho_absoluto(caminho_relativo):
    """
    Obtém o caminho absoluto a partir de um caminho relativo
    
    Args:
        caminho_relativo (str): Caminho relativo
        
    Returns:
        str: Caminho absoluto
    """
    import os
    return os.path.abspath(caminho_relativo)

def procurar_e_clicar_proximo(driver):
    """
    Procura e clica no botão 'Próximo' usando a função procurar_e_clicar melhorada
    
    Args:
        driver: WebDriver do Selenium
        
    Returns:
        bool: True se o clique foi bem-sucedido, False caso contrário
    """
    logger.info("Procurando botão 'Próximo'...")
    
    # Lista de seletores para o botão 'Próximo'
    seletores_proximo = [
        # Seletor exato do HTML problemático
        'button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_janela_proximo',
        'button.estrutura_botao_janela_proximo',
        'button[name="botao_proximo"]',
        'button[myaccesskey="p"]',
        
        # Seletores adicionais
        'button[name*="proximo"]',
        'button[id*="proximo"]',
        'button[title*="Próximo"]',
        'a[title*="Próximo"]',
        'button.botao-proximo',
        'button.btn-primary'
    ]
    
    # Utiliza a função procurar_e_clicar com texto específico
    return procurar_e_clicar(driver, seletores_proximo, texto_botao="Próximo", max_tentativas=3, espera=1)

def procurar_e_clicar(driver, seletores, texto_botao=None, max_tentativas=3, espera=1):
    """
    Procura e clica em um elemento usando uma lista de seletores e múltiplas estratégias
    
    Args:
        driver: WebDriver do Selenium
        seletores (list): Lista de seletores CSS para procurar
        texto_botao (str, opcional): Texto do botão para busca adicional por texto
        max_tentativas (int, opcional): Número máximo de tentativas para cada método
        espera (int, opcional): Tempo de espera entre tentativas em segundos
        
    Returns:
        bool: True se o clique foi bem-sucedido, False caso contrário
    """
    logger.info(f"Procurando elemento para clicar{' com texto: '+texto_botao if texto_botao else ''}...")
    
    # Adiciona seletores específicos para os botões problemáticos
    seletores_especificos = [
        # Seletor exato para o botão "Próximo"
        "button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_janela_proximo",
        "button.estrutura_botao_janela_proximo",
        "button[name='botao_proximo']",
        "button[myaccesskey='p']",
        
        # Seletor exato para o botão "Emitir"
        "button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_colorido",
        "button[name='confirmar']",
        "button[myaccesskey='e']",
        "button[disabledenableaftersubmit='enable']"
    ]
    
    # Combina seletores específicos com os seletores fornecidos
    todos_seletores = seletores_especificos + seletores
    
    # Adiciona XPaths baseados no texto do botão
    xpaths = []
    if texto_botao:
        xpaths = [
            f"//button[contains(text(), '{texto_botao}')]",
            f"//button[normalize-space()='{texto_botao}']",
            f"//input[@value='{texto_botao}']",
            f"//a[contains(text(), '{texto_botao}')]",
            f"//span[contains(text(), '{texto_botao}')]/parent::button"
        ]
    
    # Tenta cada seletor CSS fornecido
    for seletor in todos_seletores:
        try:
            # Espera pelo elemento ficar visível primeiro
            try:
                WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, seletor))
                )
                logger.info(f"Elemento ficou visível com seletor: {seletor}")
            except Exception as e:
                logger.debug(f"Elemento não ficou visível com seletor {seletor}: {e}")
                
            # Mesmo se o wait falhar, tenta localizar diretamente
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            elementos_visiveis = [e for e in elementos if e.is_displayed()]
            
            if elementos_visiveis:
                elemento = elementos_visiveis[0]
                logger.info(f"Elemento encontrado com seletor: {seletor}")
                
                # Rola até o elemento para garantir que esteja visível
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
                time.sleep(espera)
                
                # Tenta diferentes métodos de clique com várias tentativas
                metodos = [
                    {"nome": "clique direto", "func": lambda e=elemento: e.click()},
                    {"nome": "JavaScript", "func": lambda e=elemento: driver.execute_script("arguments[0].click();", e)},
                    {"nome": "ActionChains", "func": lambda e=elemento: ActionChains(driver).move_to_element(e).click().perform()},
                    {"nome": "JavaScript avançado", "func": lambda e=elemento: driver.execute_script(
                        """
                        var evt = new MouseEvent('click', {
                            bubbles: true,
                            cancelable: true,
                            view: window
                        });
                        arguments[0].dispatchEvent(evt);
                        """, 
                        e
                    )},
                    {"nome": "submit", "func": lambda e=elemento: driver.execute_script("arguments[0].form.submit();", e)},
                ]
                
                for metodo in metodos:
                    for tentativa in range(max_tentativas):
                        try:
                            logger.info(f"Tentando clicar com método: {metodo['nome']} (tentativa {tentativa+1}/{max_tentativas})")
                            metodo["func"]()
                            logger.info(f"Elemento clicado com sucesso via {metodo['nome']}")
                            
                            # Aguarda a página carregar
                            time.sleep(2)
                            esperar_pagina_carregar(driver, 5)
                            
                            # Salva screenshot para verificação
                            salvar_screenshot(driver, f"apos_clicar_elemento_{metodo['nome'].replace(' ', '_')}.png")
                            
                            # Verifica se realmente mudou de página
                            time.sleep(1)  # Espera um pouco mais para verificar
                            
                            return True
                            
                        except Exception as e:
                            logger.warning(f"Clique via {metodo['nome']} falhou na tentativa {tentativa+1}: {e}")
                            time.sleep(espera)  # Espera entre tentativas
                            
                            if tentativa == max_tentativas - 1:
                                logger.warning(f"Todas as tentativas com {metodo['nome']} falharam.")
                                continue
                        
        except Exception as e:
            logger.debug(f"Erro ao procurar elemento com seletor {seletor}: {e}")
    
    # Se os seletores CSS falharem e temos texto do botão, tenta por XPath
    if xpaths:
        for xpath in xpaths:
            try:
                elementos = driver.find_elements(By.XPATH, xpath)
                elementos_visiveis = [e for e in elementos if e.is_displayed()]
                
                if elementos_visiveis:
                    elemento = elementos_visiveis[0]
                    logger.info(f"Elemento encontrado com XPath: {xpath}")
                    
                    # Rola até o elemento
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
                    time.sleep(espera)
                    
                    try:
                        elemento.click()
                        logger.info("Elemento clicado com sucesso via XPath")
                        time.sleep(2)
                        salvar_screenshot(driver, "apos_clicar_elemento_xpath.png")
                        return True
                    except Exception as e:
                        logger.warning(f"Clique direto por XPath falhou, tentando via JavaScript: {e}")
                        try:
                            driver.execute_script("arguments[0].click();", elemento)
                            logger.info("Elemento clicado via JavaScript (XPath)")
                            time.sleep(2)
                            salvar_screenshot(driver, "apos_clicar_elemento_xpath_js.png")
                            return True
                        except Exception as e2:
                            logger.warning(f"Clique via JavaScript (XPath) falhou: {e2}")
                            
            except Exception as e:
                logger.debug(f"Erro ao procurar elemento com XPath {xpath}: {e}")
    
    logger.error("Nenhum elemento foi encontrado ou clicado com sucesso")
    salvar_screenshot(driver, "erro_elemento_nao_encontrado.png")
    return False

def extrair_numero_nota_fiscal(driver):
    """
    Extrai o número da nota fiscal da página de confirmação
    
    Args:
        driver: WebDriver do Selenium
        
    Returns:
        str: Número da nota fiscal ou None se não encontrado
    """
    logger.info("Tentando extrair número da nota fiscal...")
    
    # Lista de seletores possíveis para encontrar o número da nota
    seletores_numero = [
        # Campos de input que podem conter o número
        'input[name*="numero"]',
        'input[id*="numero"]',
        'input[name*="nota"]',
        'input[id*="nota"]',
        
        # Spans ou divs que podem exibir o número
        'span[id*="numero"]',
        'div[id*="numero"]',
        'span[class*="numero"]',
        'div[class*="numero"]',
        
        # Labels que podem mostrar o número
        'label[for*="numero"]',
        'label[class*="numero"]'
    ]
    
    # XPaths para buscar por texto
    xpaths_numero = [
        "//div[contains(text(), 'Número da nota')]//following-sibling::*",
        "//span[contains(text(), 'Número da nota')]//following-sibling::*",
        "//label[contains(text(), 'Número da nota')]//following-sibling::*",
        "//div[contains(text(), 'Número:')]//following-sibling::*",
        "//span[contains(text(), 'Número:')]//following-sibling::*",
        "//div[contains(text(), 'NFS-e:')]//following-sibling::*",
        "//span[contains(text(), 'NFS-e:')]//following-sibling::*",
        "//div[contains(text(), 'Nota Fiscal:')]//following-sibling::*",
        "//span[contains(text(), 'Nota Fiscal:')]//following-sibling::*",
        
        # Busca direta por elementos que contenham números
        "//div[contains(text(), 'Número da nota:')]/text()[normalize-space()]",
        "//span[contains(text(), 'Número da nota:')]/text()[normalize-space()]",
        "//p[contains(text(), 'Número da nota:')]/text()[normalize-space()]"
    ]
    
    # Padrões regex para extrair números
    import re
    padrao_numero = re.compile(r'\b\d{3,}\b')  # Busca por números com pelo menos 3 dígitos
    
    # Tenta encontrar através de seletores CSS
    for seletor in seletores_numero:
        try:
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor)
            for elemento in elementos:
                if elemento.is_displayed():
                    # Tenta obter o valor do elemento
                    valor = elemento.get_attribute('value') or elemento.text.strip()
                    if valor:
                        numeros_encontrados = padrao_numero.findall(valor)
                        if numeros_encontrados:
                            numero = numeros_encontrados[0]
                            logger.info(f"Número da nota encontrado via seletor {seletor}: {numero}")
                            return numero
        except Exception as e:
            logger.debug(f"Erro ao procurar número com seletor {seletor}: {e}")
    
    # Tenta encontrar através de XPaths
    for xpath in xpaths_numero:
        try:
            elementos = driver.find_elements(By.XPATH, xpath)
            for elemento in elementos:
                if elemento.is_displayed():
                    texto = elemento.text.strip()
                    if texto:
                        numeros_encontrados = padrao_numero.findall(texto)
                        if numeros_encontrados:
                            numero = numeros_encontrados[0]
                            logger.info(f"Número da nota encontrado via XPath {xpath}: {numero}")
                            return numero
        except Exception as e:
            logger.debug(f"Erro ao procurar número com XPath {xpath}: {e}")
    
    # Busca mais ampla no texto da página
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        
        # Procura por padrões específicos no texto
        padroes_busca = [
            r'Número da nota[:\s]+(\d{3,})',
            r'NFS-e[:\s]+(\d{3,})',
            r'Nota Fiscal[:\s]+(\d{3,})',
            r'Número[:\s]+(\d{3,})',
            r'emitida com sucesso[.\s]*(\d{3,})',
            r'gerada com sucesso[.\s]*(\d{3,})'
        ]
        
        for padrao in padroes_busca:
            match = re.search(padrao, body_text, re.IGNORECASE)
            if match:
                numero = match.group(1)
                logger.info(f"Número da nota encontrado no texto da página com padrão '{padrao}': {numero}")
                return numero
    
    except Exception as e:
        logger.debug(f"Erro ao buscar número no texto da página: {e}")
    
    # Se chegou aqui, não conseguiu encontrar o número
    logger.warning("Não foi possível extrair o número da nota fiscal automaticamente")
    salvar_screenshot(driver, "falha_extrair_numero_nota.png")
    
    # Salva o HTML da página para análise posterior
    try:
        html_content = driver.page_source
        with open("logs/html/pagina_confirmacao_nota.html", "w", encoding="utf-8") as f:
            f.write(html_content)
        logger.info("HTML da página de confirmação salvo em logs/html/pagina_confirmacao_nota.html")
    except Exception as e:
        logger.warning(f"Erro ao salvar HTML: {e}")
    
    return None

def main():
    # Importa módulos necessários no escopo local da função
    import os
    import platform
    import sys
    import traceback
    
    # PASSO 1: CARREGAMENTO DOS DADOS DO EXCEL
    logger.info("="*80)
    logger.info("INICIANDO AUTOMAÇÃO DE EMISSÃO DE NOTAS FISCAIS")
    logger.info("="*80)
    
    # Carrega os dados do Excel
    df_excel = carregar_dados_excel(EXCEL_PATH)
    if df_excel is None:
        logger.error("Falha ao carregar dados do Excel. Encerrando automação.")
        return

    driver = None  # Inicializa variável do driver fora do loop
    navegador_aberto = False  # Flag para controlar se o navegador já está aberto

    while True:
        # Encontra a próxima nota a ser processada
        proxima_nota = encontrar_proxima_nota(df_excel)
        if proxima_nota is None:
            logger.info("Nenhuma nota pendente encontrada. Todas as notas podem já ter sido processadas.")
            break

        # Mapeia os dados da nota
        dados_nota = mapear_dados_nota(proxima_nota['dados'])
        if dados_nota is None:
            logger.error("Falha ao mapear dados da nota. Encerrando automação.")
            continue

        linha_excel = proxima_nota['linha_excel']
        logger.info(f"Nota a ser processada encontrada na linha {linha_excel} do Excel")
        
        # Verifica o sistema operacional
        sistema_operacional = platform.system()
        logger.info(f"Sistema Operacional detectado: {sistema_operacional}")
        
        try:
            # PASSO 2: INICIANDO O NAVEGADOR
            logger.info("Configurando navegador...")
            
            # Configurações do Chrome
            chrome_opts = Options()
            chrome_opts.add_argument("--start-maximized")
            chrome_opts.add_argument("--disable-notifications")
            
            # Inicia o navegador
            driver = webdriver.Chrome(options=chrome_opts)
            logger.info("Navegador iniciado com sucesso!")
            
            # PASSO 3: NAVEGANDO PARA A PÁGINA INICIAL
            logger.info(f"Acessando URL: {NFS_URL}")
            driver.get(NFS_URL)
            
            # Aguarda carregamento da página
            esperar_pagina_carregar(driver, timeout=30)
            # Captura estado inicial
            salvar_screenshot(driver, "pagina_inicial.png")
            logger.info(f"Screenshot inicial salvo em {obter_caminho_absoluto('logs/imagens/pagina_inicial.png')}")
            
            # Log informativo
            logger.info(f"Título da página: {driver.title}")
            logger.info(f"URL atual: {driver.current_url}")
              # Salvar HTML inicial
            salvar_html(driver, "pagina_inicial")
            
            # PASSO 3: PROCESSO DE LOGIN
            logger.info("Iniciando processo de login...")
            login_sucesso = realizar_login(driver, CPF_CNPJ, SENHA)
            
            if login_sucesso:
                logger.info("LOGIN REALIZADO COM SUCESSO!")
                
                # PASSO 4: ACESSAR ÁREA FISCAL
                logger.info("Tentando acessar área fiscal...")
                acessar_sucesso = clicar_acessar_fiscal(driver)
            
            if acessar_sucesso:
                logger.info("Botão 'Acessar' clicado com sucesso!")
                
                # AVISO SOBRE CAPTCHA NESTE MOMENTO
                logger.info("\n" + "*"*80)
                logger.info("ATENÇÃO: RESOLVA O CAPTCHA AGORA!")
                logger.info("Um CAPTCHA deve aparecer após clicar no botão 'Acessar'.")
                logger.info("Por favor resolva o CAPTCHA manualmente na janela do navegador.")
                logger.info("*"*80 + "\n")
                
                # Pausa para permitir resolução manual de CAPTCHA
                input("Pressione ENTER depois de resolver o CAPTCHA para continuar...")
                
                # PASSO 5: AGUARDAR REDIRECIONAMENTO PARA A PÁGINA DE DESTINO
                logger.info(f"Aguardando redirecionamento para: {PAGINA_DESTINO}")
                logger.info("Este processo pode levar vários minutos. Por favor, aguarde...")
                  # Aguarda até 5 minutos (300 segundos) pelo redirecionamento
                destino_alcancado = aguardar_pagina_destino(driver, PAGINA_DESTINO, tempo_maximo=300)
                
                if destino_alcancado:
                    logger.info("PÁGINA DE DESTINO ALCANÇADA COM SUCESSO!")
                    # Captura estado final
                    salvar_screenshot(driver, "pagina_destino.png")
                    salvar_html(driver, "pagina_destino")
                    
                    # PASSO 6: FECHAR AVISO NA PÁGINA DE DESTINO
                    logger.info("Tentando fechar o aviso na página de destino...")
                    aviso_fechado = fechar_aviso(driver)
                    
                    if aviso_fechado:
                        logger.info("AVISO FECHADO COM SUCESSO!")
                        salvar_screenshot(driver, "apos_fechar_aviso.png")
                    else:
                        logger.warning("Não foi possível fechar o aviso automaticamente. Pode ser necessário fechá-lo manualmente.")
                    
                    logger.info("AUTENTICAÇÃO CONCLUÍDA COM SUCESSO!")
                    
                    # PASSO 7: CLICAR NO BOTÃO "EMITIR NOTA FISCAL"
                    logger.info("Tentando clicar no botão 'Emitir Nota Fiscal'...")
                    emitir_nota_sucesso = clicar_emitir_nota_fiscal(driver)
                    
                    if emitir_nota_sucesso:
                        logger.info("BOTÃO 'EMITIR NOTA FISCAL' CLICADO COM SUCESSO!")
                        logger.info("Aguardando carregamento da próxima página...")
                        time.sleep(3)  # Pausa para carregamento
                        
                        # PASSO 8: CLICAR NO BOTÃO "PRÓXIMO"
                        logger.info("Tentando clicar no botão 'Próximo'...")
                        proximo_sucesso = clicar_proximo(driver)
                        
                        if proximo_sucesso:
                            logger.info("BOTÃO 'PRÓXIMO' CLICADO COM SUCESSO!")
                            # Verificar se o fluxo de emissão foi iniciado corretamente
                            time.sleep(3)  # Pequena pausa para carregamento
                            emissao_iniciada = verificar_emissao_iniciada(driver)
                            
                            if emissao_iniciada:
                                logger.info("FLUXO DE EMISSÃO DE NOTA FISCAL INICIADO COM SUCESSO!")
                                
                                # PASSO 9: SELECIONAR TIPO DO TOMADOR
                                logger.info("Selecionando 'Pessoa Jurídica' como tipo do tomador...")
                                selecao_sucesso = selecionar_tipo_tomador(driver, tipo="Pessoa Jurídica")
                                if selecao_sucesso:
                                    logger.info("TIPO DO TOMADOR 'PESSOA JURÍDICA' SELECIONADO COM SUCESSO!")
                                    salvar_screenshot(driver, "tipo_tomador_selecionado.png")
                                    logger.info("Continuando com o preenchimento dos dados da nota fiscal...")
                                    
                                    # Buscar empresa por CNPJ usando o módulo busca_empresa
                                    try:
                                        # Importa o módulo de busca de empresa
                                        import sys
                                        import os
                                        
                                        # Garante que o diretório atual está no path
                                        if os.path.dirname(os.path.abspath(__file__)) not in sys.path:
                                            sys.path.append(os.path.dirname(os.path.abspath(__file__)))
                                        
                                        # Agora importa a função de busca de empresa
                                        from busca_empresa import preencher_busca_cnpj
                                        
                                        # Obtém os dados necessários
                                        cnpj = dados_nota.get('cnpj_tomador', '')
                                        nome_empresa = dados_nota.get('razao_social', '')
                                        
                                        # Log dos dados (mascarando o CNPJ por segurança)
                                        if len(cnpj) > 6:
                                            cnpj_mascarado = f"{cnpj[:4]}****{cnpj[-2:]}"
                                        else:
                                            cnpj_mascarado = "****"
                                        
                                        logger.info(f"Buscando empresa: {nome_empresa}")
                                        logger.info(f"CNPJ: {cnpj_mascarado}")
                                        
                                        # Salva screenshot antes da busca
                                        salvar_screenshot(driver, "antes_busca_empresa.png")
                                        
                                        # Usa a função do módulo busca_empresa para localizar e selecionar a empresa
                                        # Passa o logger para manter o registro de logs consistente
                                        busca_resultado = preencher_busca_cnpj(driver, cnpj, nome_empresa, logger)
                                        
                                        if busca_resultado:
                                            logger.info("EMPRESA ENCONTRADA E SELECIONADA COM SUCESSO!")
                                            salvar_screenshot(driver, "apos_selecao_empresa.png")
                                        else:
                                            logger.warning("Não foi possível encontrar ou selecionar a empresa pelo CNPJ automaticamente.")
                                            logger.warning("Tentando abordagem alternativa...")
                                            
                                            # Pergunta ao usuário se deseja continuar manualmente
                                            continuar_manual = input("Empresa não encontrada automaticamente. Deseja selecionar manualmente? (s/n): ")
                                            if continuar_manual.lower() == 's':
                                                logger.info("Aguardando seleção manual da empresa pelo usuário...")
                                                input("Selecione a empresa manualmente e pressione ENTER para continuar...")
                                                logger.info("Continuando após seleção manual da empresa")
                                            else:
                                                logger.warning("Processo interrompido pelo usuário após falha na busca por empresa")
                                                return False
                                    except ImportError as e:
                                        logger.error(f"Erro ao importar módulo busca_empresa.py: {e}")
                                        logger.error("Verifique se o arquivo busca_empresa.py está presente no diretório")
                                        return False
                                    except Exception as e:
                                        logger.error(f"Erro durante a busca de empresa: {e}")
                                        import traceback
                                        logger.error(traceback.format_exc())
                                        
                                        # Pergunta ao usuário se deseja continuar manualmente
                                        continuar_manual = input("Erro durante a busca de empresa. Deseja continuar manualmente? (s/n): ")
                                        if continuar_manual.lower() == 's':
                                            logger.info("Continuando manualmente após erro...")
                                        else:
                                            logger.warning("Processo interrompido pelo usuário após erro")
                                            return False
                                    
                                    # PASSO 10: PREENCHER DADOS DO TOMADOR
                                    logger.info("Preenchendo dados do tomador com informações do Excel...")
                                    logger.info(f"CNPJ Tomador: {dados_nota.get('cnpj_tomador', '')[:4]}****{dados_nota.get('cnpj_tomador', '')[-2:]}")
                                    logger.info(f"Razão Social: {dados_nota.get('razao_social', '')}")
                                    preenchimento_tomador = preencher_dados_tomador(driver, dados_nota)
                                    
                                    if preenchimento_tomador:
                                        logger.info("DADOS DO TOMADOR PREENCHIDOS COM SUCESSO!")
                                           # PASSO 11: PREENCHER DADOS DO SERVIÇO
                                        logger.info("Preenchendo dados do serviço...")
                                        try:
                                            # Tenta importar o módulo especializado
                                            from preencher_dados_servico import preencher_formulario_servico
                                            
                                            # Usa a função do módulo especializado passando o logger
                                            preenchimento_servico = preencher_formulario_servico(driver, dados_nota, logger)
                                        except ImportError:
                                            logger.warning("Módulo preencher_dados_servico não encontrado, usando função interna")
                                            preenchimento_servico = preencher_dados_servico(driver, dados_nota)
                                        except Exception as e:
                                            logger.error(f"Erro ao usar módulo especializado: {e}")
                                            logger.warning("Tentando com a função interna...")
                                            preenchimento_servico = preencher_dados_servico(driver, dados_nota)
                                        
                                        if preenchimento_servico:
                                            logger.info("DADOS DO SERVIÇO PREENCHIDOS COM SUCESSO!")
                                            
                                            # PASSO 12: AVANÇAR NO FLUXO
                                            logger.info("Tentando avançar para a próxima etapa...")
                                            avancar_sucesso = procurar_e_clicar_proximo(driver)
                                            
                                            if avancar_sucesso:
                                                logger.info("AVANÇOU COM SUCESSO PARA A PRÓXIMA ETAPA!")
                                                salvar_screenshot(driver, "formulario_preenchido_avancar.png")
                                                
                                                # PASSO 13: PREENCHER TRIBUTOS FEDERAIS
                                                logger.info("Preenchendo tributos federais (IR, PIS, COFINS, CSLL)...")
                                                try:
                                                    # Tenta importar o módulo especializado para tributos
                                                    from preencher_tributos import preencher_tributos
                                                    
                                                    # Verifica se há tributos definidos na nota
                                                    tem_tributos = any([
                                                        dados_nota.get('irrf', dados_nota.get('valor_ir', 0)) != 0,
                                                        dados_nota.get('pis', dados_nota.get('valor_pis', 0)) != 0,
                                                        dados_nota.get('cofins', dados_nota.get('valor_cofins', 0)) != 0,
                                                        dados_nota.get('csll', dados_nota.get('valor_csll', 0)) != 0
                                                    ])
                                                    
                                                    # Se tem tributos, preenche os campos
                                                    if tem_tributos:
                                                        logger.info("Preenchendo campos de tributos federais...")
                                                        # Usa a função do módulo especializado passando o logger
                                                        preenchimento_tributos = preencher_tributos(driver, dados_nota, logger)
                                                        
                                                        if preenchimento_tributos:
                                                            logger.info("TRIBUTOS FEDERAIS PREENCHIDOS COM SUCESSO!")
                                                        else:
                                                            logger.warning("Houve problemas ao preencher tributos federais")
                                                            # Pergunta se deseja continuar mesmo assim
                                                            continuar_tributos = input("Houve problemas ao preencher os tributos federais. Deseja continuar mesmo assim? (s/n): ")
                                                            if continuar_tributos.lower() != 's':
                                                                logger.warning("Processo interrompido pelo usuário após falha nos tributos federais")
                                                                return
                                                    else:
                                                        logger.info("Nota não possui tributos federais para preencher")
                                                except ImportError:
                                                    logger.warning("Módulo preencher_tributos não encontrado, pulando essa etapa")
                                                except Exception as e:
                                                    logger.error(f"Erro ao preencher tributos federais: {e}")
                                                    import traceback
                                                    logger.error(traceback.format_exc())
                                                    # Pergunta se deseja continuar mesmo assim
                                                    continuar_tributos = input("Erro ao preencher os tributos federais. Deseja continuar mesmo assim? (s/n): ")
                                                    if continuar_tributos.lower() != 's':
                                                        logger.warning("Processo interrompido pelo usuário após erro nos tributos federais")
                                                        return
                                                
                                                salvar_screenshot(driver, "apos_preencher_tributos.png")
                                                
                                                # PASSO 14: FINALIZAR EMISSÃO                                                
                                                logger.info("FORMULÁRIO PREENCHIDO - PRONTO PARA EMISSÃO")
                                                logger.info("Verifique manualmente se todos os dados estão corretos")
                                                
                                                # Pergunta ao usuário se deseja continuar com a emissão
                                                continuar = input("Todos os dados estão corretos? Pressione ENTER para emitir a nota ou 'n' para cancelar: ")
                                                if continuar.lower() != 'n':
                                                    # PASSO 13: Finalizar emissão da nota fiscal
                                                    logger.info("Finalizando a emissão da nota fiscal...")
                                                          # Tenta clicar no botão para emitir a nota
                                                    # Lista de seletores possíveis para o botão Emitir
                                                    seletores_emitir = [
                                                        # Seletor exato do HTML problemático
                                                        "button.__estrutura_componente_base.botao.botao-com-variante.estrutura_botao.disabled_user_select.estrutura_botao_colorido",
                                                        "button[name='confirmar']",
                                                        "button[myaccesskey='e']",
                                                        "button[disabledenableaftersubmit='enable']",
                                                        
                                                        # Seletores alternativos
                                                        "button[name='emitir']", 
                                                        "button.botao-primario",
                                                        "button.__estrutura_componente_base.botao.botao-primario", 
                                                        "button[type='submit']"
                                                    ]
                                                    
                                                    # Tenta clicar usando a função procurar_e_clicar com texto específico
                                                    emitir_sucesso = procurar_e_clicar(driver, seletores_emitir, texto_botao="Emitir", max_tentativas=3, espera=1)
                                                    
                                                    if emitir_sucesso:
                                                        logger.info("NOTA FISCAL EMITIDA COM SUCESSO!")
                                                        salvar_screenshot(driver, "nota_emitida.png")
                                                        
                                                        # Aguarda um tempo para ter certeza que a página de confirmação carregou
                                                        logger.info("Aguardando carregamento da página de confirmação...")
                                                        time.sleep(5)
                                                        
                                                        # Extrai o número da nota fiscal
                                                        logger.info("Tentando extrair o número da nota fiscal...")
                                                        numero_nota = extrair_numero_nota_fiscal(driver)
                                                        
                                                        if numero_nota:
                                                            logger.info(f"NÚMERO DA NOTA EXTRAÍDO COM SUCESSO: {numero_nota}")
                                                            
                                                            # Atualiza o número da nota no Excel
                                                            logger.info(f"Atualizando Excel com o número da nota: {numero_nota}")
                                                            atualizacao_sucesso = atualizar_numero_nota_excel(
                                                                EXCEL_PATH, linha_excel, numero_nota)
                                                            
                                                            if atualizacao_sucesso:
                                                                logger.info("ARQUIVO EXCEL ATUALIZADO COM SUCESSO!")
                                                            else:
                                                                logger.error("FALHA AO ATUALIZAR O ARQUIVO EXCEL")
                                                                logger.error(f"Por favor, atualize manualmente o número da nota {numero_nota} na linha {linha_excel} do Excel")
                                                        else:
                                                            logger.warning("Não foi possível extrair o número da nota automaticamente")
                                                            numero_manual = input("Por favor, informe o número da nota fiscal emitida (deixe em branco para ignorar): ")
                                                            if numero_manual.strip():
                                                                logger.info(f"Atualizando Excel com o número informado manualmente: {numero_manual}")
                                                                atualizacao_sucesso = atualizar_numero_nota_excel(
                                                                    EXCEL_PATH, linha_excel, numero_manual)
                                                                if atualizacao_sucesso:
                                                                    logger.info("ARQUIVO EXCEL ATUALIZADO COM SUCESSO!")
                                                                else:
                                                                    logger.error("FALHA AO ATUALIZAR O ARQUIVO EXCEL")
                                                            else:
                                                                logger.warning("Nenhum número informado. O Excel não será atualizado.")
                                                    else:
                                                        logger.warning("Não foi possível clicar no botão para finalizar a emissão")
                                                        logger.info("Prossiga manualmente para completar a emissão da nota fiscal")
                                                else:
                                                    logger.info("Emissão de nota fiscal cancelada pelo usuário")
                                                
                                                logger.info("PROCESSO DE AUTOMAÇÃO CONCLUÍDO PARA ESTA NOTA")
                                                
                                                # Pergunta se deseja continuar com próxima nota
                                                continuar_proxima = input("Deseja processar a próxima nota? (s/n): ")
                                                if continuar_proxima.lower() != 's':
                                                    logger.info("Processo interrompido pelo usuário")
                                                    break
                                            else:
                                                logger.warning("Não foi possível avançar automaticamente")
                                                logger.info("Formulário preenchido - continue manualmente")                                                # Em caso de erro, pergunta se quer tentar a próxima nota
                                                tentar_proxima = input("Houve erro nesta nota. Deseja tentar a próxima? (s/n): ")
                                                if tentar_proxima.lower() != 's':
                                                    break
                                                    
        except Exception as e:
            logger.error(f"Erro durante a automação: {e}")
            import traceback
            logger.error(traceback.format_exc())
            
            if 'driver' in locals():
                salvar_screenshot(driver, "erro_execucao.png")
                logger.info("Screenshot do erro salvo em logs/imagens/erro_execucao.png")
                salvar_html(driver, "pagina_erro")
            
            logger.error("\nSUGESTÕES PARA RESOLVER O PROBLEMA:")
            logger.error("1. Certifique-se de que o arquivo .env existe e contém as variáveis necessárias")
            logger.error("2. Verifique se sua conexão com a internet está estável")
            logger.error("3. Tente acessar o site manualmente para confirmar que está funcionando")
              # Em caso de erro, pergunta se quer tentar a próxima nota
            tentar_proxima = input("Houve erro nesta nota. Deseja tentar a próxima? (s/n): ")
            if tentar_proxima.lower() != 's':
                break
                
        finally:
            # Verifica se podemos continuar com a próxima nota sem fechar o navegador
            continuar_sem_fechar = False
            
            if 'driver' in locals() and 'emitir_sucesso' in locals() and emitir_sucesso:
                try:
                    # Importa o módulo fora do bloco if para evitar problemas de indentação
                    import importlib.util
                    spec = importlib.util.spec_from_file_location("detectar_continuar_emissao", "c:/nfe/detectar_continuar_emissao.py")
                    modulo_detectar = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(modulo_detectar)
                    
                    logger.info("Verificando se podemos continuar automaticamente para a próxima nota...")
                    continuacao_automatica = modulo_detectar.continuar_emissao_apos_nota(driver)
                    
                    if continuacao_automatica:
                        logger.info("Continuando automaticamente para a próxima nota sem fechar o navegador")
                        continuar_sem_fechar = True
                except Exception as e:
                    logger.error(f"Erro ao tentar continuar para próxima nota: {e}")
                    continuacao_automatica = False
            
            # Se não foi possível continuar automaticamente, fecha o navegador
            if not continuar_sem_fechar and 'driver' in locals():
                logger.info("Fechando o navegador após processar nota")
                driver.quit()
                logger.info("Navegador fechado")
                
        # Recarrega os dados do Excel para capturar possíveis atualizações
        df_excel = carregar_dados_excel(EXCEL_PATH)

# O fluxo completo de automação implementado é:
#
# 1. Inicialização e configuração do navegador
# 2. Acesso à página inicial do sistema NFSe
# 3. Login no sistema com CPF/CNPJ e senha
# 4. Clicar no botão "Acessar" para entrar na área fiscal
# 5. Aguardar redirecionamento para a página de destino (pode levar vários minutos)
# 6. Fechar o aviso na página de destino, se houver
# 7. Clicar no botão "Emitir Nota Fiscal" para iniciar o processo de emissão
# 8. Clicar no botão "Próximo" para avançar no fluxo de emissão
# 9. Selecionar "Pessoa Jurídica" como tipo do tomador
#
# Cada etapa possui tratamento de erros e capturas de tela para facilitar a depuração.
# Screenshots são salvos na pasta logs/imagens com nomes descritivos.

if __name__ == "__main__":
    main()